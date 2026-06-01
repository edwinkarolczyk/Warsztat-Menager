"""Tests for the independent Plan Monitor engine."""

import json
from datetime import datetime

import pandas as pd
from openpyxl import Workbook

from PlanMonitor.main import (PlanMonitor, compare_plans, export_changes,
                              load_config, normalize_quantity, parse_plan,
                              position_key, read_plan, resolve_mapping,
                              safe_date, save_config)


def row(order="306", symbol="1.670.93 REMS", quantity=336,
        date="2026-06-11", process=""):
    return {"order": order, "symbol": symbol, "quantity": quantity,
            "date": date, "process": process}


def make_realistic_plan(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plan produkcji"
    sheet.append(["PLAN PRODUKCJI", None, None, "Czerwiec"])
    sheet.append(["Nr zlec.", None, "Ilość szt", "Data wysyłki", "Proces",
                  "Tydzień", "Pon", "Wt"])
    sheet.append([None, None, None, None, None, "24", "10", "11"])
    sheet.append([306, "00.055 REMS", "450 szt", "10 cze", "zgrzane"])
    sheet.append([None, "1.670.93 REMS - RAL1003", 336, None, "lakiernia"])
    sheet.append([None, None, None, None, None])
    sheet.append([None, "2.510.120 REMS", 20, None, "pakowanie"])
    sheet.append([None, "Pon", None, None, None])
    sheet.append([500, "NT detal", 12, datetime(2026, 6, 12), "produkcja"])
    workbook.save(path)


def test_parse_plan_scans_complete_sheet_and_inherits_group_values(tmp_path):
    plan = tmp_path / "plan.xlsx"
    make_realistic_plan(plan)

    parsed = parse_plan(plan)

    assert parsed.rows == [
        row(symbol="00.055 REMS", quantity=450, date="10 cze", process="zgrzane"),
        row(symbol="1.670.93 REMS - RAL1003", date="10 cze", process="lakiernia"),
        row(symbol="2.510.120 REMS", quantity=20, date="10 cze", process="pakowanie"),
        row("500", "NT detal", 12, "2026-06-12", "produkcja"),
    ]
    diagnostics = parsed.diagnostics
    assert diagnostics.rows_scanned == 9
    assert diagnostics.records_count == 4
    assert diagnostics.skipped_empty == 1
    assert diagnostics.skipped_header == 1
    assert diagnostics.skipped_without_symbol >= 1
    assert diagnostics.column_mapping == {
        "order": "A", "symbol": "B", "quantity": "C", "date": "D",
        "process": "E",
    }
    assert len(diagnostics.sample) == 4


def test_blank_row_does_not_stop_parser(tmp_path):
    plan = tmp_path / "plan.xlsx"
    make_realistic_plan(plan)

    assert read_plan(plan)[-1]["order"] == "500"


def test_manual_mapping_and_selected_sheet_override_detection(tmp_path):
    plan = tmp_path / "plan.xlsx"
    workbook = Workbook()
    workbook.active.title = "Nie ten"
    sheet = workbook.create_sheet("Właściwy")
    sheet.append(["opis bez klasycznego nagłówka", "x", "x", "x", "x"])
    sheet.append(["REMS symbol", 700, "10 cze", "gotowe", 900])
    workbook.save(plan)

    parsed = parse_plan(plan, {"sheet_name": "Właściwy", "data_start_row": 2,
                               "column_mapping": {"symbol": "A", "quantity": "B",
                                                  "date": "C", "process": "D",
                                                  "order": "E"}})

    assert parsed.rows == [row("900", "REMS symbol", 700, "10 cze", "gotowe")]


def test_safe_date_and_quantity_normalizers_handle_empty_values():
    assert safe_date(None) == ""
    assert safe_date(pd.NaT) == ""
    assert safe_date(float("nan")) == ""
    assert safe_date(" 10\ncze ") == "10 cze"
    assert safe_date(datetime(2026, 6, 10)) == "2026-06-10"
    assert normalize_quantity(450.0) == 450
    assert normalize_quantity("450 szt") == 450
    assert normalize_quantity("") is None
    assert normalize_quantity("brak") is None


def test_compare_uses_stable_key_and_detects_quantity_date_and_process():
    previous = [row(process="cięcie")]
    current = [row(quantity=450, date="2026-06-13", process="pakowanie")]

    changes = compare_plans(previous, current)

    assert {change["type"] for change in changes} == {
        "quantity_changed", "date_changed", "process_changed",
    }
    assert not {"new", "removed"} & {change["type"] for change in changes}


def test_symbol_change_is_reported_as_removed_and_new():
    changes = compare_plans([row()], [row(symbol="NOWY REMS")])

    assert {change["type"] for change in changes} == {"new", "removed"}


def test_no_order_key_includes_symbol_and_date():
    assert position_key(row(order="", date="10 cze")) == (
        "NO_ORDER|1.670.93 REMS|10 cze"
    )


def test_check_saves_all_records_and_parser_metadata_without_wm_output(tmp_path):
    config_path = tmp_path / "config.json"
    snapshot_path = tmp_path / "snapshots" / "current_snapshot.json"
    history_path = tmp_path / "reports" / "history.jsonl"
    plan_file = tmp_path / "plan.xlsx"
    make_realistic_plan(plan_file)
    config = load_config(tmp_path / "missing.json")
    config["plan_file"] = str(plan_file)
    save_config(config, config_path)
    monitor = PlanMonitor(config_path, snapshot_path, history_path)

    result = monitor.check(force=True)

    assert result.status == "changed"
    snapshot = json.loads(snapshot_path.read_text(encoding="utf-8"))
    assert snapshot["source_file"] == str(plan_file)
    assert snapshot["sheet"] == "Plan produkcji"
    assert snapshot["parser"]["records_count"] == 4
    assert len(snapshot["records"]) == 4
    assert not (tmp_path / "data" / "pending_dispositions.json").exists()


def test_export_csv_and_txt_include_parser_summary(tmp_path):
    changes = compare_plans([], [row()], ["REMS"], "2026-06-01T07:30:00")
    parser = {"rows_scanned": 124, "records_count": 87,
              "column_mapping": {"order": "A", "symbol": "B", "quantity": "C",
                                 "date": "D", "process": "E"}}

    export_changes(changes, tmp_path / "report.csv", parser)
    export_changes(changes, tmp_path / "report.txt", parser)

    csv_text = (tmp_path / "report.csv").read_text(encoding="utf-8-sig")
    txt_text = (tmp_path / "report.txt").read_text(encoding="utf-8")
    assert "Przeskanowano wierszy: 124" in csv_text
    assert "Znaleziono pozycji: 87" in txt_text
    assert "Użyte kolumny: A/B/C/D/E" in txt_text
    assert "DOTYCZY DZIAŁU" in csv_text
    assert "NOWE" in txt_text


def test_resolve_mapping_accepts_aliases_and_manual_letters():
    columns = ["Nr zlecenia", "bez nagłówka", "Ilość", "Termin", "Operacja"]

    assert resolve_mapping(columns, {"symbol": "B"}) == {
        "order": "A", "symbol": "B", "quantity": "C", "date": "D",
        "process": "E",
    }
