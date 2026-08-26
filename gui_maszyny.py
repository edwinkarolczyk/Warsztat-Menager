# version: 1.6
# Zmiany 1.6:
# - Karty/PDF maszyn zapisują się w centralnym ROOT WM: <root>/wydruki/karty, niezależnie od katalogu uruchomienia.
# Zmiany 1.5:
# - Cykliczny przegląd ma dokładny dzień miesiąca (domyślnie 1) i nadal powtarza się co roku.
# - Serwis cykliczny i jego automatyczna Dyspozycja synchronizują rozpoczęcie oraz wykonanie.
# - Najbliższy przegląd i zaległość są czytelniejsze na głównej liście; dodano proste ostrzeżenie o terminach.
# - Historia serwisowa pokazuje powiązaną Dyspozycję i pełniejsze szczegóły wpisu.
# - Usunięto stare podpięcie wm.dyspo_wizard; przycisk Maszyn otwiera aktywny kreator Dyspozycji.
# - Harmonogram pomocniczy używa bieżącego roku zamiast stałego 2025.
# - Aktywne okno Użytkowanie maszyny używa formatu Start/Stop: Dzień DD-MM-RR HH:MM.
# Zmiany 1.4:
# - Start/Stop w historii statusów pokazują polski skrót dnia tygodnia oraz datę DD-MM-RR z godziną.
# Zmiany 1.3:
# - Cykliczne miesiące przeglądów są widoczne w dolnej liście serwisów jako Przegląd cykliczny.
# - Rozpoczęcie lub wykonanie wpisu cyklicznego materializuje go do reviews bez duplikowania miesiąca.
# Zmiany 1.2:
# - Uproszczono okno Użytkowanie maszyny: mniej kolumn w historii i serwisach.
# - Historia statusów pokazuje bezpośrednio powód / zdarzenie serwisowe.
# - Historia statusów odświeża się od razu po rozpoczęciu i zakończeniu serwisu.
# Zmiany 1.1:
# - Główna lista Maszyn używa tej samej wielkości czcionki co Dyspozycje: Segoe UI 11, nagłówki 11 bold, wiersz 30 px.
from __future__ import annotations

import calendar
import datetime as dt
import os
import shutil
import subprocess
import tkinter as tk
from logging import getLogger
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple

from core.settings_manager import Settings
from gui_maszyny_view import MachinesView

try:
    from PIL import Image, ImageTk  # type: ignore
except Exception:  # pragma: no cover - PIL opcjonalne
    Image = None
    ImageTk = None

def _maybe_open_dyspo(root, context):
    """Otwórz aktywny kreator Dyspozycji z kontekstem Maszyn."""

    try:
        from gui_dyspozycje_creator import open_dyspozycje_creator
    except Exception as exc:
        messagebox.showerror(
            "Maszyny",
            f"Nie udało się otworzyć kreatora Dyspozycji:\n{exc}",
            parent=root if getattr(root, "tk", None) is not None else None,
        )
        return

    target = root
    if hasattr(root, "winfo_toplevel"):
        try:
            target = root.winfo_toplevel()
        except Exception:
            target = root

    ctx = {
        "typ_dyspozycji": "maszyna",
        "modul_zrodlowy": "maszyny",
    }
    if isinstance(context, dict):
        ctx.update(context)
    try:
        open_dyspozycje_creator(
            target,
            autor=_active_login_for_machine(target),
            context=ctx,
        )
    except Exception as exc:
        messagebox.showerror(
            "Maszyny",
            f"Nie udało się otworzyć kreatora Dyspozycji:\n{exc}",
            parent=target if getattr(target, "tk", None) is not None else None,
        )


CANVAS_W = 1000
CANVAS_H = 1000

GRID_BASE_BG_PX_X = 25
GRID_BASE_BG_PX_Y = 25

DEFAULT_BG_COLOR = "#1e1e1e"

MACHINE_STATUS_ALIASES = {
    "ok": "ok",
    "sprawna": "ok",
    "sprawne": "ok",
    "sprawny": "ok",
    "dziala": "ok",
    "działa": "ok",
    "alert": "alert",
    "serwis": "alert",
    "przeglad": "alert",
    "przegląd": "alert",
    "serwis/przeglad": "alert",
    "serwis/przegląd": "alert",
    "warn": "warn",
    "warm": "warn",
    "warning": "warn",
    "awaria": "warn",
    "uszkodzona": "warn",
    "uszkodzone": "warn",
    "stop": "warn",
}

MACHINE_STATUS_LABELS = {
    "ok": "Sprawna",
    "alert": "Serwis / przegląd",
    "warn": "Awaria",
}

MACHINE_STATUS_COLORS = {
    "ok": "#16a34a",
    "alert": "#ca8a04",
    "warn": "#dc2626",
}

MACHINE_STATUS_ROW_COLORS = {
    "ok": {"background": "#dcfce7", "foreground": "#166534"},
    "alert": {"background": "#fef3c7", "foreground": "#854d0e"},
    "warn": {"background": "#fee2e2", "foreground": "#7f1d1d"},
}

MACHINE_STATUS_EDIT_LABELS = {
    "Sprawna": "ok",
    "Serwis / przegląd": "alert",
    "Awaria": "warn",
}

MACHINE_STATUS_EDIT_VALUES = list(MACHINE_STATUS_EDIT_LABELS.keys())


SCHEDULE_YEAR = dt.date.today().year
SCHEDULE_SOON_THRESHOLD_DAYS = 7
SCHEDULE_STATUS_COLORS = {
    "overdue": MACHINE_STATUS_COLORS["warn"],
    "soon": MACHINE_STATUS_COLORS["alert"],
    "ok": MACHINE_STATUS_COLORS["ok"],
    "done": "#0f766e",
    "none": "#64748b",
}
SCHEDULE_STATUS_ROW_COLORS = {
    "overdue": MACHINE_STATUS_ROW_COLORS["warn"],
    "soon": MACHINE_STATUS_ROW_COLORS["alert"],
    "ok": MACHINE_STATUS_ROW_COLORS["ok"],
    "done": {"background": "#e0f2fe", "foreground": "#0c4a6e"},
    "none": {"background": "#e2e8f0", "foreground": "#475569"},
}
_TREE_STATUS_TAG_CACHE: Dict[str, bool] = {}

from core.path_utils import resolve_root_path
from utils_json import (
    normalize_doc_list_or_dict,
    safe_read_json as _safe_read_json,
    safe_write_json as _safe_write_json,
)


def _safe_clamp(v: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, v))


def _canvas_bounds(canvas) -> tuple[int, int]:
    try:
        w = int(canvas.winfo_width())
        h = int(canvas.winfo_height())
        return max(1, w), max(1, h)
    except Exception:
        return (800, 600)

logger = getLogger(__name__)


def _normalize_machine_status(value: object) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return "ok"
    key = raw.replace("_", " ").replace("-", " ")
    key = " ".join(key.split())
    direct = MACHINE_STATUS_ALIASES.get(key)
    if direct:
        return direct
    compact = key.replace(" ", "")
    return MACHINE_STATUS_ALIASES.get(compact, key)


def _machine_status_label(value: object) -> str:
    key = _normalize_machine_status(value)
    return MACHINE_STATUS_LABELS.get(key, str(value or "Sprawna"))


def _machine_status_edit_label(value: object) -> str:
    key = _normalize_machine_status(value)
    return MACHINE_STATUS_LABELS.get(key, "Sprawna")


MONTH_LABELS_PL = [
    (1, "Styczeń"),
    (2, "Luty"),
    (3, "Marzec"),
    (4, "Kwiecień"),
    (5, "Maj"),
    (6, "Czerwiec"),
    (7, "Lipiec"),
    (8, "Sierpień"),
    (9, "Wrzesień"),
    (10, "Październik"),
    (11, "Listopad"),
    (12, "Grudzień"),
]


def _normalize_review_months(value: object) -> List[int]:
    if isinstance(value, list):
        raw_values = value
    elif value in (None, ""):
        raw_values = []
    else:
        raw_values = [value]
    out: List[int] = []
    for item in raw_values:
        try:
            month = int(item)
        except Exception:
            continue
        if 1 <= month <= 12 and month not in out:
            out.append(month)
    return sorted(out)


def _machine_now_iso() -> str:
    return dt.datetime.now().replace(microsecond=0).isoformat()


def _parse_machine_dt(value: object) -> Optional[dt.datetime]:
    if isinstance(value, dt.datetime):
        return value
    if not value:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.endswith("Z"):
        text = text[:-1]
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return dt.datetime.strptime(text[:19], fmt)
        except ValueError:
            pass
    try:
        return dt.datetime.fromisoformat(text)
    except Exception:
        return None


_MACHINE_WEEKDAY_LABELS_PL = ("Pon", "Wt", "Śr", "Czw", "Pt", "Sob", "Nie")


def _format_machine_history_dt(value: object) -> str:
    parsed = _parse_machine_dt(value)
    if parsed is None:
        raw = str(value or "").strip()
        return raw.replace("T", " ")[:16] if raw else "—"
    weekday = _MACHINE_WEEKDAY_LABELS_PL[parsed.weekday()]
    return f"{weekday} {parsed.strftime('%d-%m-%y %H:%M')}"


def _duration_minutes(started_at: object, ended_at: object) -> int:
    start = _parse_machine_dt(started_at)
    end = _parse_machine_dt(ended_at)
    if not start or not end:
        return 0
    return max(0, int((end - start).total_seconds())) // 60


def _format_duration_minutes(minutes: object) -> str:
    try:
        total = int(minutes or 0)
    except Exception:
        total = 0
    if total < 60:
        return f"{total} min"
    hours, mins = divmod(total, 60)
    if hours < 24:
        return f"{hours}h {mins}m"
    days, hours = divmod(hours, 24)
    if days < 30:
        return f"{days}d {hours}h"
    months, days = divmod(days, 30)
    return f"{months} mies. {days}d"


def _active_login_for_machine(root: tk.Misc | None = None) -> str:
    candidates = [root]
    try:
        if root is not None and hasattr(root, "winfo_toplevel"):
            candidates.append(root.winfo_toplevel())
    except Exception:
        pass
    for source in candidates:
        if source is None:
            continue
        for attr in ("active_login", "_wm_login", "login"):
            value = str(getattr(source, attr, "") or "").strip()
            if value:
                return value
    return "system"


def _ensure_status_current(machine: Dict[str, Any], *, actor: str = "system") -> None:
    if isinstance(machine.get("status_current"), dict):
        return
    status = _normalize_machine_status(machine.get("status"))
    machine["status_current"] = {
        "status": status,
        "label": _machine_status_label(status),
        "started_at": _machine_now_iso(),
        "changed_by": actor,
        "note": "",
        "photos": [],
    }


def _apply_machine_status_change(
    machine: Dict[str, Any],
    new_status: str,
    *,
    actor: str,
    note: str,
    photos: Optional[List[str]] = None,
) -> bool:
    photos = list(photos or [])
    old_status = _normalize_machine_status(machine.get("status"))
    new_status = _normalize_machine_status(new_status)
    if old_status == new_status:
        machine["status"] = new_status
        _ensure_status_current(machine, actor=actor)
        return False

    now = _machine_now_iso()
    history = machine.get("status_history")
    if not isinstance(history, list):
        history = []
        machine["status_history"] = history

    current = machine.get("status_current")
    if not isinstance(current, dict):
        current = {
            "status": old_status,
            "label": _machine_status_label(old_status),
            "started_at": now,
            "changed_by": actor,
            "note": "",
            "photos": [],
        }

    closed = dict(current)
    closed.setdefault("status", old_status)
    closed.setdefault("label", _machine_status_label(old_status))
    closed["ended_at"] = now
    closed["duration_minutes"] = _duration_minutes(closed.get("started_at"), now)
    closed["closed_by"] = actor
    closed["close_note"] = note or ""
    history.append(closed)

    machine["status"] = new_status
    machine["status_current"] = {
        "status": new_status,
        "label": _machine_status_label(new_status),
        "started_at": now,
        "changed_by": actor,
        "note": note or "",
        "photos": photos,
    }
    return True


def _machine_attachment_root() -> str:
    """Zwraca katalog ROOT/data/maszyny/attachments."""

    try:
        from core import root_paths as wm_root_paths

        data_root = wm_root_paths.get_data_root()
        if data_root:
            return os.path.join(str(data_root), "maszyny", "attachments")
    except Exception:
        pass

    try:
        from config_manager import ConfigManager

        data_root = ConfigManager().path_data()
        if data_root:
            return os.path.join(str(data_root), "maszyny", "attachments")
    except Exception:
        pass

    return os.path.join("data", "maszyny", "attachments")


def _safe_machine_file_part(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        text = "machine"
    out = []
    for ch in text:
        if ch.isalnum() or ch in ("-", "_"):
            out.append(ch)
        else:
            out.append("_")
    return "".join(out).strip("_") or "machine"


def _copy_machine_status_photos(
    machine_id: object, source_paths: Iterable[str]
) -> List[str]:
    """Kopiuje zdjęcia statusu do ROOT/data/maszyny/attachments."""

    copied: List[str] = []
    valid_sources = [
        str(path) for path in source_paths or [] if str(path or "").strip()
    ]
    if not valid_sources:
        return copied

    ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    machine_part = _safe_machine_file_part(machine_id)
    target_dir = os.path.join(_machine_attachment_root(), machine_part, ts)
    os.makedirs(target_dir, exist_ok=True)

    for idx, src in enumerate(valid_sources, start=1):
        if not os.path.exists(src):
            continue
        ext = os.path.splitext(src)[1].lower() or ".jpg"
        filename = f"status_{idx:02d}{ext}"
        dst = os.path.join(target_dir, filename)
        try:
            shutil.copy2(src, dst)
            copied.append(dst)
        except Exception:
            logger.exception(
                "[MASZYNY][STATUS_PHOTO] Nie udało się skopiować zdjęcia: %s",
                src,
            )
    return copied


REVIEW_TYPES = (
    "Przegląd okresowy",
    "Serwis planowany",
    "Konserwacja",
    "Kalibracja",
    "Czyszczenie",
    "Inne",
)

REVIEW_STATUS_PLANNED = "planned"
REVIEW_STATUS_DONE = "done"
REVIEW_STATUS_CANCELLED = "cancelled"

REVIEW_SOURCE_CYCLE = "cycle"
REVIEW_SOURCE_MANUAL = "manual"

REVIEW_SOURCE_LABELS = {
    REVIEW_SOURCE_CYCLE: "Cykliczny",
    REVIEW_SOURCE_MANUAL: "Ręczny",
}


def _machine_reviews(machine: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Zwraca lokalne przeglądy zapisane bezpośrednio w rekordzie maszyny."""

    reviews = machine.get("reviews")
    if isinstance(reviews, list):
        return [item for item in reviews if isinstance(item, dict)]
    return []


def _machine_default_review_type(machine: Dict[str, Any]) -> str:
    value = (
        machine.get("default_review_type")
        or machine.get("domyslny_typ_przegladu")
        or machine.get("typ_przegladu")
        or "Przegląd okresowy"
    )
    text = str(value or "").strip()
    return text if text else "Przegląd okresowy"


def _machine_suggested_service_people(machine: Dict[str, Any]) -> List[str]:
    value = (
        machine.get("suggested_service_people")
        or machine.get("serwisanci")
        or machine.get("wykonawcy_serwis")
        or machine.get("wykonawcy")
        or []
    )
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item or "").strip()]
    return _split_csv_people(value)


def _review_status_key(value: object) -> str:
    raw = str(value or "").strip().lower()
    raw = raw.replace("_", " ").replace("-", " ")
    raw = " ".join(raw.split())
    if raw in (
        "done",
        "wykonany",
        "wykonane",
        "zrobione",
        "zamkniety",
        "zamknięty",
        "completed",
    ):
        return REVIEW_STATUS_DONE
    if raw in ("cancelled", "canceled", "anulowany", "anulowane"):
        return REVIEW_STATUS_CANCELLED
    return REVIEW_STATUS_PLANNED


def _review_date(value: object) -> Optional[dt.date]:
    return _parse_schedule_date(value)


def _format_machine_review_date(value: object) -> str:
    parsed = _review_date(value)
    if parsed is None:
        raw = str(value or "").strip()
        return raw[:10] if raw else "—"
    weekday = _MACHINE_WEEKDAY_LABELS_PL[parsed.weekday()]
    return f"{weekday} {parsed.strftime('%d-%m-%y')}"


def _machine_review_months(machine: Dict[str, Any]) -> List[int]:
    value = (
        machine.get("review_months")
        or machine.get("inspection_months")
        or machine.get("miesiace_przegladu")
        or machine.get("miesiące_przeglądu")
        or machine.get("months")
        or []
    )
    return _normalize_review_months(value)


def _machine_review_day(machine: Dict[str, Any]) -> int:
    try:
        day = int(machine.get("review_day") or machine.get("inspection_day") or 1)
    except (TypeError, ValueError):
        day = 1
    return max(1, min(31, day))


def _cycle_review_date(year: int, month: int, day: int) -> dt.date:
    last_day = calendar.monthrange(int(year), int(month))[1]
    return dt.date(int(year), int(month), min(max(1, int(day)), last_day))


def _review_month_done(
    machine: Dict[str, Any],
    *,
    year: int,
    month: int,
    review_type: str,
) -> bool:
    wanted_type = str(review_type or "").strip().lower()
    for review in _machine_reviews(machine):
        if _review_status_key(review.get("status")) != REVIEW_STATUS_DONE:
            continue
        date_value = _review_date(
            review.get("date")
            or review.get("data")
            or review.get("planned_date")
            or review.get("completed_at")
            or review.get("done_at")
        )
        if not date_value:
            continue
        if date_value.year != year or date_value.month != month:
            continue
        current_type = str(review.get("type") or review.get("typ") or "").strip().lower()
        if not current_type or current_type == wanted_type:
            return True
    return False


def _review_month_has_persisted_cycle(
    machine: Dict[str, Any],
    *,
    year: int,
    month: int,
    review_type: str,
) -> bool:
    wanted_type = str(review_type or "").strip().lower()
    for review in _machine_reviews(machine):
        if str(review.get("source") or "").strip().lower() != REVIEW_SOURCE_CYCLE:
            continue
        date_value = _review_date(
            review.get("date")
            or review.get("data")
            or review.get("planned_date")
            or review.get("completed_at")
            or review.get("done_at")
        )
        if not date_value or date_value.year != year or date_value.month != month:
            continue
        current_type = str(review.get("type") or review.get("typ") or "").strip().lower()
        if not current_type or current_type == wanted_type:
            return True
    return False


def _combined_machine_review_entries(
    machine: Dict[str, Any],
    *,
    today: Optional[dt.date] = None,
    years_ahead: int = 1,
) -> List[Dict[str, Any]]:
    """
    Wspólna lista harmonogramu maszyny:
    - ręczne wpisy machine["reviews"],
    - planowane wpisy cykliczne wygenerowane z miesięcy przeglądu.
    """

    today = today or dt.date.today()
    entries: List[Dict[str, Any]] = []

    for review in _machine_reviews(machine):
        date_value = _review_date(
            review.get("date")
            or review.get("data")
            or review.get("planned_date")
            or review.get("completed_at")
            or review.get("done_at")
        )
        if date_value is None:
            continue

        entry = dict(review)
        entry["date"] = date_value.isoformat()
        entry["type"] = str(review.get("type") or review.get("typ") or "Przegląd okresowy")
        entry["status"] = str(review.get("status") or REVIEW_STATUS_PLANNED)
        entry["source"] = str(review.get("source") or REVIEW_SOURCE_MANUAL)
        entries.append(entry)

    default_type = _machine_default_review_type(machine)
    suggested_people = _machine_suggested_service_people(machine)
    months = _machine_review_months(machine)
    years = range(today.year, today.year + max(1, years_ahead) + 1)

    for year in years:
        for month in months:
            if _review_month_has_persisted_cycle(
                machine, year=year, month=month, review_type=default_type
            ):
                continue
            if _review_month_done(machine, year=year, month=month, review_type=default_type):
                continue

            planned_date = _cycle_review_date(
                year, month, _machine_review_day(machine)
            )
            entries.append(
                {
                    "id": f"cycle_{year}_{month:02d}",
                    "date": planned_date.isoformat(),
                    "type": default_type,
                    "status": REVIEW_STATUS_PLANNED,
                    "source": REVIEW_SOURCE_CYCLE,
                    "suggested_people": suggested_people,
                }
            )

    entries.sort(
        key=lambda item: (
            _review_date(item.get("date")) or dt.date(9999, 12, 31),
            str(item.get("type") or ""),
            str(item.get("source") or ""),
        )
    )
    return entries


def _next_machine_review_entry(
    machine: Dict[str, Any],
    *,
    today: Optional[dt.date] = None,
) -> Optional[Dict[str, Any]]:
    today = today or dt.date.today()
    planned: List[Dict[str, Any]] = []

    for entry in _combined_machine_review_entries(machine, today=today):
        if _review_status_key(entry.get("status")) != REVIEW_STATUS_PLANNED:
            continue
        if _review_date(entry.get("date")) is None:
            continue
        planned.append(entry)

    if not planned:
        return None

    return min(
        planned,
        key=lambda item: _review_date(item.get("date")) or dt.date(9999, 12, 31),
    )


def _combined_machine_schedule_summary(machine: Dict[str, Any]) -> Dict[str, Any]:
    today = dt.date.today()
    entries = _combined_machine_review_entries(machine, today=today)
    entry = _next_machine_review_entry(machine, today=today)

    history = [
        item
        for item in entries
        if _review_status_key(item.get("status")) == REVIEW_STATUS_DONE
    ]
    upcoming = [
        item
        for item in entries
        if _review_status_key(item.get("status")) == REVIEW_STATUS_PLANNED
    ]

    if not entry:
        return {
            "upcoming": upcoming,
            "history": history,
            "next_entry": None,
            "next_date": None,
            "status": "none",
            "key": "none",
            "status_key": "none",
            "next_label": "—",
            "status_label": "Brak danych",
            "status_text": "Brak zaplanowanych przeglądów",
            "days": None,
            "color": SCHEDULE_STATUS_COLORS["none"],
        }

    date_value = _review_date(entry.get("date"))
    if date_value is None:
        return {
            "upcoming": upcoming,
            "history": history,
            "next_entry": None,
            "next_date": None,
            "status": "none",
            "key": "none",
            "status_key": "none",
            "next_label": "—",
            "status_label": "Brak danych",
            "status_text": "Brak danych harmonogramu",
            "days": None,
            "color": SCHEDULE_STATUS_COLORS["none"],
        }

    days = (date_value - today).days
    if days < 0:
        status = "overdue"
        label = f"Po terminie • {abs(days)} dni"
    elif days == 0:
        status = "soon"
        label = "Dzisiaj"
    elif days == 1:
        status = "soon"
        label = "Jutro"
    elif days <= SCHEDULE_SOON_THRESHOLD_DAYS:
        status = "soon"
        label = f"Za {days} dni"
    else:
        status = "ok"
        label = f"Za {days} dni"

    type_label = str(entry.get("type") or "Przegląd okresowy")
    source_label = REVIEW_SOURCE_LABELS.get(
        str(entry.get("source") or ""),
        str(entry.get("source") or ""),
    )
    date_label = _format_machine_review_date(date_value)
    details = [type_label]
    if source_label:
        details.append(source_label)
    details.append(label)

    return {
        "upcoming": upcoming,
        "history": history,
        "next_entry": entry,
        "next_date": date_value,
        "status": status,
        "key": status,
        "status_key": status,
        "next_label": date_label,
        "status_label": label,
        "status_text": f"{' • '.join(details)} – {date_label}",
        "days": days,
        "color": SCHEDULE_STATUS_COLORS.get(status, SCHEDULE_STATUS_COLORS["none"]),
    }


def _new_review_id() -> str:
    return "rev_" + dt.datetime.now().strftime("%Y%m%d_%H%M%S")


def _split_csv_people(value: object) -> List[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def _load_wm_user_logins() -> List[str]:
    """Czyta loginy użytkowników WM z ROOT-a. Fallback: pusta lista."""

    candidates: List[str] = []
    try:
        from core import root_paths as wm_root_paths

        data_root = wm_root_paths.get_data_root()
        if data_root:
            candidates.extend(
                [
                    os.path.join(str(data_root), "profiles.json"),
                    os.path.join(str(data_root), "uzytkownicy.json"),
                    os.path.join(str(data_root), "users.json"),
                    os.path.join(str(data_root), "profile", "profiles.json"),
                ]
            )
    except Exception:
        pass

    try:
        from config_manager import ConfigManager

        data_root = ConfigManager().path_data()
        if data_root:
            candidates.extend(
                [
                    os.path.join(str(data_root), "profiles.json"),
                    os.path.join(str(data_root), "uzytkownicy.json"),
                    os.path.join(str(data_root), "users.json"),
                    os.path.join(str(data_root), "profile", "profiles.json"),
                ]
            )
    except Exception:
        pass

    logins: List[str] = []
    for path in candidates:
        if not path or not os.path.exists(path):
            continue
        payload = _safe_read_json(path, default={})
        raw_items = []
        if isinstance(payload, list):
            raw_items = payload
        elif isinstance(payload, dict):
            for key in ("users", "uzytkownicy", "profiles", "items"):
                value = payload.get(key)
                if isinstance(value, list):
                    raw_items = value
                    break
            if not raw_items:
                raw_items = list(payload.values())

        for item in raw_items:
            if isinstance(item, dict):
                login = str(
                    item.get("login")
                    or item.get("username")
                    or item.get("user")
                    or item.get("name")
                    or item.get("nazwa")
                    or ""
                ).strip()
            else:
                login = str(item or "").strip()
            if login and login not in logins:
                logins.append(login)
    return logins


def _machine_status_history_rows(machine: Dict[str, Any]) -> List[tuple]:
    rows: List[tuple] = []
    history = machine.get("status_history")
    if isinstance(history, list):
        for item in history:
            if not isinstance(item, dict):
                continue
            status = _machine_status_label(item.get("status"))
            start = _format_machine_history_dt(item.get("started_at"))
            stop = _format_machine_history_dt(item.get("ended_at"))
            duration = _format_duration_minutes(item.get("duration_minutes"))
            who = str(item.get("closed_by") or item.get("changed_by") or "—")
            note = str(item.get("close_note") or item.get("note") or "")
            photos = (
                item.get("photos") if isinstance(item.get("photos"), list) else []
            )
            photo_txt = f" | zdjęcia: {len(photos)}" if photos else ""
            rows.append((status, start, stop, duration, who, note + photo_txt))

    current = machine.get("status_current")
    if isinstance(current, dict):
        start_raw = current.get("started_at")
        now = _machine_now_iso()
        rows.append(
            (
                _machine_status_label(current.get("status")),
                _format_machine_history_dt(start_raw),
                "w toku",
                _format_duration_minutes(_duration_minutes(start_raw, now)),
                str(current.get("changed_by") or "—"),
                str(current.get("note") or "")
                + (
                    f" | zdjęcia: {len(current.get('photos') or [])}"
                    if isinstance(current.get("photos"), list)
                    and current.get("photos")
                    else ""
                ),
            )
        )
    return rows


def _machine_status_color(value: object) -> str:
    key = _normalize_machine_status(value)
    return MACHINE_STATUS_COLORS.get(key, MACHINE_STATUS_COLORS["ok"])


def _machine_status_row_colors(value: object) -> dict[str, str]:
    key = _normalize_machine_status(value)
    return MACHINE_STATUS_ROW_COLORS.get(key, MACHINE_STATUS_ROW_COLORS["ok"])


def _normalize_machine_key(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    filtered = [ch for ch in text if ch.isalnum()]
    normalized = "".join(filtered)
    if normalized.isdigit():
        normalized = normalized.lstrip("0") or "0"
    return normalized


def _machine_identifiers(machine: Dict[str, Any]) -> List[str]:
    identifiers: set[str] = set()
    for key in ("id", "nr_ewid", "nr", "nazwa"):
        normalized = _normalize_machine_key(machine.get(key))
        if normalized:
            identifiers.add(normalized)
    return list(identifiers)


def _entry_identifiers(entry: Dict[str, Any]) -> List[str]:
    identifiers: set[str] = set()
    for key in ("machine_id", "id", "machine", "machine_name", "nazwa"):
        normalized = _normalize_machine_key(entry.get(key))
        if normalized:
            identifiers.add(normalized)
    return list(identifiers)


def _match_schedule_entry(machine: Dict[str, Any], entry: Dict[str, Any]) -> bool:
    if not isinstance(machine, dict) or not isinstance(entry, dict):
        return False
    machine_keys = set(_machine_identifiers(machine))
    entry_keys = set(_entry_identifiers(entry))
    return bool(machine_keys & entry_keys)


def _parse_schedule_date(value: object) -> Optional[dt.date]:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        ordinal = int(value)
        if ordinal > 59:  # Excel 1900 date system offset
            try:
                return dt.date.fromordinal(ordinal + 693594)
            except ValueError:
                pass
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in (
            "%Y-%m-%d",
            "%d.%m.%Y",
            "%Y.%m.%d",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%m-%Y",
            "%Y-%m-%d %H:%M:%S",
        ):
            try:
                return dt.datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
        try:
            return dt.date.fromisoformat(raw)
        except ValueError:
            return None
    return None


def _format_next_label(entry: Dict[str, Any], date_obj: dt.date) -> str:
    label = date_obj.isoformat()
    typ = str(entry.get("type") or entry.get("typ") or "").strip()
    if typ:
        label = f"{label} ({typ})"
    return label


def _normalize_schedule_entry(raw: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not isinstance(raw, dict):
        return None

    machine_id = str(raw.get("machine_id") or raw.get("id") or "").strip()
    machine_name = str(
        raw.get("machine_name")
        or raw.get("machine")
        or raw.get("nazwa")
        or raw.get("maszyna")
        or ""
    ).strip()
    date_value = raw.get("date") or raw.get("termin") or raw.get("data")
    date_obj = _parse_schedule_date(date_value)
    if date_obj is None:
        return None

    inspection_type = str(raw.get("type") or raw.get("typ") or raw.get("rodzaj") or "").strip()
    responsible = str(
        raw.get("responsible")
        or raw.get("osoba")
        or raw.get("odpowiedzialny")
        or ""
    ).strip()
    notes = str(raw.get("notes") or raw.get("uwagi") or raw.get("komentarz") or "").strip()

    status_raw = str(raw.get("status") or "").strip().lower()
    status = "wykonany" if status_raw in {"wykonany", "done", "completed"} else "planowany"

    entry: Dict[str, Any] = {
        "machine_id": machine_id,
        "machine_name": machine_name or machine_id,
        "date": date_obj.isoformat(),
        "type": inspection_type,
        "responsible": responsible,
        "notes": notes,
        "status": status,
    }

    completed_value = raw.get("completed_at") or raw.get("wykonano")
    completed_date = _parse_schedule_date(completed_value) if completed_value else None
    if status == "wykonany":
        if completed_date is not None:
            entry["completed_at"] = completed_date.isoformat()
        elif isinstance(completed_value, str) and completed_value.strip():
            entry["completed_at"] = completed_value.strip()

    card_value = str(raw.get("card") or raw.get("karta") or "").strip()
    if card_value:
        entry["card"] = card_value

    return entry


def _schedule_entry_keys(entry: Dict[str, Any]) -> List[Tuple[str, str, str]]:
    keys: List[Tuple[str, str, str]] = []
    machine_keys = _entry_identifiers(entry)
    if not machine_keys:
        machine_keys = [""]
    date_key = str(entry.get("date") or "")
    type_key = str(entry.get("type") or "").strip().lower()
    for mk in machine_keys:
        keys.append((mk, date_key, type_key))
    return keys


def _merge_schedule_status(
    new_entries: List[Dict[str, Any]], existing_entries: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    lookup: Dict[Tuple[str, str, str], Dict[str, Any]] = {}
    for entry in existing_entries or []:
        for key in _schedule_entry_keys(entry):
            lookup[key] = entry

    seen: set[Tuple[str, str, str]] = set()
    for entry in new_entries:
        for key in _schedule_entry_keys(entry):
            seen.add(key)
            previous = lookup.get(key)
            if not previous:
                continue
            prev_status = str(previous.get("status") or "").strip().lower()
            if prev_status == "wykonany":
                entry["status"] = "wykonany"
                if previous.get("completed_at"):
                    entry["completed_at"] = previous["completed_at"]
            if previous.get("card") and not entry.get("card"):
                entry["card"] = previous["card"]

    extras: List[Dict[str, Any]] = []
    for entry in existing_entries or []:
        keys = _schedule_entry_keys(entry)
        if not keys:
            continue
        if all(key not in seen for key in keys):
            if str(entry.get("status") or "").strip().lower() == "wykonany":
                extras.append(dict(entry))

    merged = list(new_entries) + extras
    merged.sort(key=lambda item: (str(item.get("machine_id") or item.get("machine_name") or ""), str(item.get("date") or "")))
    return merged


def _schedule_summary(
    entries: Iterable[Dict[str, Any]],
    *,
    today: Optional[dt.date] = None,
    soon_threshold: int = SCHEDULE_SOON_THRESHOLD_DAYS,
) -> Dict[str, Any]:
    today = today or dt.date.today()
    parsed: List[Tuple[dt.date, Dict[str, Any]]] = []
    for entry in entries or []:
        date_obj = _parse_schedule_date(entry.get("date"))
        if date_obj is None:
            continue
        parsed.append((date_obj, entry))
    parsed.sort(key=lambda item: item[0])

    upcoming = [(date_obj, entry) for date_obj, entry in parsed if str(entry.get("status") or "").strip().lower() != "wykonany"]
    history = [
        (date_obj, entry)
        for date_obj, entry in parsed
        if str(entry.get("status") or "").strip().lower() == "wykonany"
    ]
    history.sort(key=lambda item: item[0], reverse=True)

    summary: Dict[str, Any] = {
        "upcoming": [entry for _, entry in upcoming],
        "history": [entry for _, entry in history],
        "next_entry": None,
        "next_date": None,
        "next_label": "—",
        "days": None,
        "status_key": "none",
        "status_label": "Brak danych",
        "status_text": "Brak zaplanowanych przeglądów",
        "color": SCHEDULE_STATUS_COLORS["none"],
    }

    if upcoming:
        next_date, next_entry = upcoming[0]
        days = (next_date - today).days
        if days < 0:
            status_key = "overdue"
            status_label = "Po terminie"
        elif days <= soon_threshold:
            status_key = "soon"
            status_label = "Wkrótce"
        else:
            status_key = "ok"
            status_label = "Planowany"
        summary.update(
            {
                "next_entry": next_entry,
                "next_date": next_date,
                "next_label": _format_next_label(next_entry, next_date),
                "days": days,
                "status_key": status_key,
                "status_label": status_label,
                "status_text": f"{status_label} – {summary['next_label']}",
                "color": SCHEDULE_STATUS_COLORS[status_key],
            }
        )
    elif history:
        last_date, last_entry = history[0]
        summary.update(
            {
                "next_entry": None,
                "next_date": None,
                "next_label": "—",
                "days": None,
                "status_key": "done",
                "status_label": "Wykonane",
                "status_text": f"Wykonano {last_date.isoformat()}",
                "color": SCHEDULE_STATUS_COLORS["done"],
            }
        )
    return summary


def _attach_schedule(
    rows: Iterable[Dict[str, Any]],
    schedule_entries: List[Dict[str, Any]],
    *,
    today: Optional[dt.date] = None,
    soon_threshold: int = SCHEDULE_SOON_THRESHOLD_DAYS,
) -> None:
    today = today or dt.date.today()
    for machine in rows or []:
        if not isinstance(machine, dict):
            continue
        matching = [entry for entry in schedule_entries if _match_schedule_entry(machine, entry)]
        machine["__schedule_entries"] = matching
        machine["__schedule_summary"] = _schedule_summary(
            matching,
            today=today,
            soon_threshold=soon_threshold,
        )


def _strip_schedule_fields(machine: Dict[str, Any]) -> Dict[str, Any]:
    return {k: v for k, v in machine.items() if not k.startswith("__schedule")}


def _schedule_status_key(machine: Dict[str, Any]) -> str:
    if not isinstance(machine, dict):
        return "none"
    summary = machine.get("__schedule_summary")
    if not isinstance(summary, dict):
        summary = _combined_machine_schedule_summary(machine)
        machine["__schedule_summary"] = summary
    return str(
        summary.get("status")
        or summary.get("key")
        or summary.get("status_key")
        or "none"
    )


def _ensure_tree_schedule_tag(tree: ttk.Treeview, status_key: str) -> str:
    tag = f"SCHEDULE::{status_key}"
    if tag in _TREE_STATUS_TAG_CACHE:
        return tag
    colors = SCHEDULE_STATUS_ROW_COLORS.get(status_key)
    if colors:
        tree.tag_configure(tag, **colors)
    _TREE_STATUS_TAG_CACHE[tag] = True
    return tag


def _ensure_tree_machine_status_tag(tree: ttk.Treeview, status_value: object) -> str:
    status_key = _normalize_machine_status(status_value)
    tag = f"MACHINE_STATUS::{status_key}"
    if tag in _TREE_STATUS_TAG_CACHE:
        return tag
    colors = _machine_status_row_colors(status_key)
    if colors:
        tree.tag_configure(tag, **colors)
    _TREE_STATUS_TAG_CACHE[tag] = True
    return tag


def _describe_entry_status(
    entry: Dict[str, Any], *, today: Optional[dt.date] = None
) -> Tuple[str, str]:
    today = today or dt.date.today()
    status = str(entry.get("status") or "").strip().lower()
    date_obj = _parse_schedule_date(entry.get("date"))
    if status == "wykonany":
        label = "Wykonany"
        if date_obj:
            label = f"Wykonany {date_obj.isoformat()}"
        return label, "done"
    if date_obj is None:
        return "Brak daty", "none"
    days = (date_obj - today).days
    if days < 0:
        return f"Po terminie ({abs(days)} dni)", "overdue"
    if days <= SCHEDULE_SOON_THRESHOLD_DAYS:
        return f"Wkrótce ({days} dni)", "soon"
    return f"Planowany ({days} dni)", "ok"


def _serialize_schedule_entries(entries: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    payload: List[Dict[str, Any]] = []
    for entry in entries or []:
        if not isinstance(entry, dict):
            continue
        clean: Dict[str, Any] = {
            "machine_id": entry.get("machine_id", ""),
            "machine_name": entry.get("machine_name", ""),
            "date": entry.get("date", ""),
            "type": entry.get("type", ""),
            "responsible": entry.get("responsible", ""),
            "notes": entry.get("notes", ""),
            "status": entry.get("status", "planowany"),
        }
        if entry.get("completed_at"):
            clean["completed_at"] = entry.get("completed_at")
        if entry.get("card"):
            clean["card"] = entry.get("card")
        payload.append(clean)
    return payload


def _resolve_card_storage(path: str, cfg: Any | None = None) -> str:
    normalized = os.path.normpath(path)
    try:
        if cfg is None:
            from config_manager import ConfigManager

            cfg = ConfigManager()
        data_root = cfg.path_data()
        project_root = cfg.path_root()
    except Exception:
        data_root = resolve_root_path("<root>", "data")
        project_root = resolve_root_path("<root>", "")

    for base, prefix in (
        (data_root, os.path.join("<root>", "data")),
        (project_root, "<root>"),
    ):
        try:
            if os.path.commonpath([normalized, base]) == os.path.normpath(base):
                rel = os.path.relpath(normalized, base)
                return os.path.join(prefix, rel).replace(os.sep, "/")
        except Exception:
            continue
    return normalized.replace(os.sep, "/")


def _resolve_card_absolute(path: str, cfg: Any | None = None) -> str:
    if not path:
        return ""
    try:
        if cfg is None:
            from config_manager import ConfigManager

            cfg = ConfigManager()
        root = cfg.path_root()
    except Exception:
        root = os.getcwd()
    return resolve_root_path(root, path)


def _open_external(path: str) -> bool:
    try:
        if os.name == "nt":
            os.startfile(path)  # type: ignore[attr-defined]
            return True
    except Exception:
        return False
    try:
        subprocess.Popen(["xdg-open", path])
        return True
    except Exception:
        try:
            subprocess.Popen(["open", path])
            return True
        except Exception:
            return False


_EXCEL_FIELD_ALIASES: Dict[str, Tuple[str, ...]] = {
    "machine_id": ("id", "nr", "nr ewid", "nr maszyny", "identyfikator"),
    "machine_name": ("maszyna", "nazwa", "nazwa maszyny"),
    "date": ("data", "termin", "termin przegl", "plan"),
    "type": ("typ", "rodzaj", "przegl", "rodzaj przegl"),
    "responsible": ("osoba", "odpowiedzial", "odpowiedzialny"),
    "notes": ("uwagi", "komentarz", "notat"),
    "status": ("status", "stan"),
    "completed_at": ("wykonano", "data wykonania", "zrealizowano"),
    "card": ("karta", "plik", "załącznik", "zalacznik"),
}


def _normalize_excel_header(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("\n", " ").replace("\xa0", " ")
    return " ".join(text.split())


def _map_excel_headers(headers: List[str]) -> List[Optional[str]]:
    mapped: List[Optional[str]] = []
    for header in headers:
        normalized = _normalize_excel_header(header)
        field: Optional[str] = None
        for key, tokens in _EXCEL_FIELD_ALIASES.items():
            if any(token in normalized for token in tokens):
                field = key
                break
        mapped.append(field)
    return mapped


def _read_excel_schedule(path: str) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError(
            "Brak biblioteki openpyxl – zainstaluj ją aby importować harmonogram."
        ) from exc

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [_normalize_excel_header(cell) for cell in header_row]
    mapping = _map_excel_headers(headers)

    entries: List[Dict[str, Any]] = []
    for row in rows_iter:
        if not row:
            continue
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        raw: Dict[str, Any] = {}
        for idx, cell in enumerate(row):
            field = mapping[idx] if idx < len(mapping) else None
            if not field:
                continue
            if field in {"machine_id", "machine_name", "type", "responsible", "notes", "status", "card"}:
                raw[field] = str(cell or "").strip()
            else:
                raw[field] = cell
        normalized = _normalize_schedule_entry(raw)
        if normalized:
            entries.append(normalized)
    return entries


def _import_schedule_from_excel(path: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    entries = _read_excel_schedule(path)
    if not entries:
        raise ValueError("Nie znaleziono danych harmonogramu w pliku Excel.")

    years = {
        _parse_schedule_date(entry.get("date")).year
        for entry in entries
        if _parse_schedule_date(entry.get("date")) is not None
    }
    year = next(iter(years)) if len(years) == 1 else SCHEDULE_YEAR
    meta = {
        "year": int(year),
        "source": os.path.basename(path),
        "imported_at": dt.datetime.now().isoformat(),
    }
    return entries, meta
def _coalesce_data_root(cfg: dict | None = None) -> str:
    """Return an absolute data root derived from *cfg* or environment."""

    cfg = cfg or {}
    paths_cfg = cfg.get("paths") or {}
    candidates = (
        paths_cfg.get("data_root"),
        cfg.get("data_root"),
        os.environ.get("WM_DATA_ROOT"),
    )
    for candidate in candidates:
        if isinstance(candidate, str) and candidate.strip():
            value = os.path.expanduser(candidate.strip())
            if not os.path.isabs(value):
                value = os.path.normpath(os.path.join(os.getcwd(), value))
            return os.path.normpath(value)
    return os.path.normpath(os.path.join(os.getcwd(), "data"))


# ---- bezpieczny import config_manager / fallback (jak w Twojej bazie) ----
try:
    from config_manager import get_config, resolve_rel
except Exception:
    def resolve_rel(cfg: dict, what: str) -> str:
        root = _coalesce_data_root(cfg)
        mapping = {"machines": os.path.join("maszyny", "maszyny.json")}
        rel = mapping.get(what)
        if not rel:
            return root
        return os.path.normpath(os.path.join(root, rel))

    def get_config() -> dict:
        try:
            from config_manager import ConfigManager  # type: ignore
            return ConfigManager().load()
        except Exception:
            return {}

from ui_theme import ensure_theme_applied

# ---- funkcje z utils_maszyny (masz po R-06Jb) ----
from utils_maszyny import (
    load_machines_rows_with_fallback,
    ensure_machines_sample_if_empty,
    load_machines,
    upsert_machine,
    delete_machine,
    merge_rows_union_by_id,
    resolve_schedule_path,
)
from maszyny_dyspozycje import (
    find_cycle_dyspozycja_for_review,
    sync_review_to_dyspozycja,
)


def load_machines_rows() -> list[dict]:
    """Wczytaj rekordy maszyn z aktywnego źródła danych aplikacji."""

    cfg = get_config() or {}
    rows, _primary_path = load_machines_rows_with_fallback(cfg, resolve_rel)
    if rows:
        return rows
    fallback_rows, _mode, _primary_count, _legacy_count = load_machines()
    return fallback_rows


def _iter_inspection_dates(machine: dict) -> list[dt.date]:
    """
    Zwraca listę poprawnych dat (datetime.date) z machine['zadania'].
    Ignoruje rekordy bez 'data' lub o błędnym formacie.
    """

    dates: list[dt.date] = []
    schedule_entries = machine.get("__schedule_entries") if isinstance(machine, dict) else None
    if schedule_entries:
        for entry in schedule_entries:
            status = str(entry.get("status") or "").strip().lower()
            if status == "wykonany":
                continue
            parsed = _parse_schedule_date(entry.get("date"))
            if parsed is not None:
                dates.append(parsed)
        return dates

    try:
        tasks = machine.get("zadania", []) or []
        for t in tasks:
            dstr = (t or {}).get("data")
            if not dstr or not isinstance(dstr, str):
                continue
            try:
                if len(dstr) == 7 and dstr.count("-") == 1:
                    d = dt.datetime.strptime(f"{dstr}-01", "%Y-%m-%d").date()
                else:
                    d = dt.datetime.strptime(dstr, "%Y-%m-%d").date()
                dates.append(d)
            except Exception:
                continue
    except Exception:
        logger.debug("[Maszyny] _iter_inspection_dates: ignoruję błędy formatu")
    return dates


def _next_inspection_date_safe(
    machine: dict, today: dt.date | None = None
) -> dt.date | None:
    """
    Zwraca NAJBLIŻSZĄ PRZYSZŁĄ datę przeglądu lub None, gdy nie istnieje.
    Nigdy nie rzuca ValueError przy pustej liście.
    """

    today = today or dt.date.today()
    dates = _iter_inspection_dates(machine)
    if not dates:
        return None
    future = [d for d in dates if d >= today]
    if not future:
        return None
    try:
        return min(future)
    except Exception:
        return None


def _days_to_next_inspection_safe(
    machine: dict, today: dt.date | None = None
) -> int | None:
    """
    Różnica dni do najbliższego przyszłego przeglądu; None gdy brak przyszłych terminów.
    """

    today = today or dt.date.today()
    nxt = _next_inspection_date_safe(machine, today=today)
    if nxt is None:
        return None
    try:
        return (nxt - today).days
    except Exception:
        return None


def _next_inspection_date(
    machine: dict, today: dt.date | None = None
) -> Optional[dt.date]:
    return _next_inspection_date_safe(machine, today=today)


def _days_to_next_inspection(machine: dict, today: dt.date | None = None) -> Optional[int]:
    return _days_to_next_inspection_safe(machine, today=today)


def _status_color(machine: dict) -> str:
    """Zwraca kolor statusu z uwzględnieniem liczby dni do przeglądu."""

    status = (machine.get("status") or "").strip().lower()
    if status in ("awaria", "uszkodzona", "error"):
        return "#dc2626"  # red-600

    summary = machine.get("__schedule_summary") if isinstance(machine, dict) else None
    if summary and summary.get("color"):
        return str(summary["color"])

    if status in ("ok", "sprawna", "sprawne"):
        return "#16a34a"  # green-600
    if status in ("uwaga", "serwis", "warning"):
        return "#ca8a04"  # yellow-600

    days = _days_to_next_inspection(machine)
    if days is None:
        return "#64748b"  # slate-500 – brak danych
    if days < 0:
        return "#dc2626"
    if days <= 30:
        return "#ca8a04"
    return "#16a34a"


def _map_label_text(machine: dict, label_mode: str) -> str:
    label_mode = (label_mode or "id").lower()
    if label_mode == "typ":
        return str(machine.get("typ", "") or "")
    if label_mode == "nazwa":
        return str(machine.get("nazwa", "") or "")
    return str(machine.get("id") or machine.get("nr_ewid") or "")


def _render_days_label_on_canvas(
    canvas, x: int, y: int, machine: dict
) -> Optional[int]:
    """
    Rysuje etykietę „dni do przeglądu” pod kropką maszyny.
    - '—' (neutralny szary) gdy brak przyszłych terminów (tylko przeszłe albo brak zadań).
    - czerwony, gdy days < 0 (spóźnione).
    """

    days_value = _days_to_next_inspection_safe(machine)
    if days_value is None:
        days_label = "—"
        fill_color = "#d1d5db"
    else:
        days_label = f"{days_value} dni"
        fill_color = "#d1d5db" if days_value >= 0 else "#fca5a5"
    try:
        return canvas.create_text(
            x, y, text=days_label, fill=fill_color, font=("Segoe UI", 8)
        )
    except Exception:
        logger.debug("[Maszyny] canvas text draw skipped (timing)")
        return None


class ImageTooltip:
    def __init__(self, parent: tk.Misc):
        self.parent = parent
        self.top: Optional[tk.Toplevel] = None
        self._img_ref = None

    def show(self, x: int, y: int, machine: dict) -> None:
        self.hide()
        self.top = tk.Toplevel(self.parent)
        self.top.wm_overrideredirect(True)
        self.top.attributes("-topmost", True)
        self.top.configure(bg="#111214")
        self.top.geometry(f"+{x + 12}+{y + 12}")

        path = machine.get("image") or machine.get("obraz") or ""
        name = machine.get("nazwa") or ""
        ttk.Label(self.top, text=name or "(bez nazwy)").pack(fill="x", padx=6, pady=(6, 2))

        if Image and ImageTk and path and os.path.exists(path):
            try:
                image = Image.open(path)
                image.thumbnail((256, 256))
                photo = ImageTk.PhotoImage(image)
                self._img_ref = photo
                tk.Label(self.top, image=photo, bd=0).pack(padx=6, pady=(0, 6))
            except Exception:
                ttk.Label(self.top, text="Nie można wczytać obrazu").pack(padx=6, pady=(0, 6))
        else:
            ttk.Label(self.top, text="Brak obrazu").pack(padx=6, pady=(0, 6))

        days = _days_to_next_inspection(machine)
        info = (
            f"Typ: {machine.get('typ', '') or '-'}   •   "
            f"Dni do przeglądu: {days if days is not None else '—'}"
        )
        ttk.Label(self.top, text=info).pack(padx=6, pady=(0, 6))

    def hide(self) -> None:
        if self.top is not None:
            self.top.destroy()
            self.top = None
            self._img_ref = None


class MonthYearDialog(tk.Toplevel):
    def __init__(self, master, title="Wybierz miesiąc i rok", init_year=None, init_month=None):
        super().__init__(master)
        self.title(title)
        self.resizable(False, False)
        self.result: Optional[str] = None
        yr = init_year or tk.IntVar(value=int(self._now("%Y")))
        mo = init_month or tk.StringVar(value=self._now("%m"))
        self._yr, self._mo = yr, mo

        body = ttk.Frame(self, padding=12)
        body.pack(fill="both", expand=True)

        ttk.Label(body, text="Rok:").grid(row=0, column=0, sticky="w", padx=(0, 6))
        self.spn_year = ttk.Spinbox(body, from_=2000, to=2100, width=6, textvariable=yr)
        self.spn_year.grid(row=0, column=1, sticky="w")

        ttk.Label(body, text="Miesiąc:").grid(row=1, column=0, sticky="w", padx=(0, 6), pady=(8, 0))
        self.cbo_month = ttk.Combobox(
            body,
            width=9,
            state="readonly",
            values=[f"{i:02d}" for i in range(1, 13)],
            textvariable=mo,
        )
        self.cbo_month.grid(row=1, column=1, sticky="w", pady=(8, 0))

        btns = ttk.Frame(body)
        btns.grid(row=2, column=0, columnspan=2, pady=(12, 0), sticky="e")
        ttk.Button(btns, text="Anuluj", command=self._on_cancel).pack(side="right")
        ttk.Button(btns, text="OK", command=self._on_ok).pack(side="right", padx=(0, 8))

        self.grab_set()
        self.transient(master)
        self.cbo_month.focus_set()

    def _now(self, fmt: str) -> str:
        import datetime as _dt

        return _dt.datetime.now().strftime(fmt)

    def _on_ok(self):
        y = int(self._yr.get())
        m = int(self._mo.get())
        self.result = f"{y:04d}-{m:02d}-01"
        self.destroy()

    def _on_cancel(self):
        self.result = None
        self.destroy()


def pick_machine_image(parent) -> Optional[str]:
    path = filedialog.askopenfilename(
        parent=parent,
        title="Wybierz zdjęcie maszyny",
        filetypes=[
            ("Obrazy", "*.jpg;*.jpeg;*.png;*.bmp;*.gif"),
            ("Wszystkie pliki", "*.*"),
        ],
    )
    return path or None


def _add_inspection_date(parent, machine: dict):
    dlg = MonthYearDialog(parent)
    parent.wait_window(dlg)
    if dlg.result:
        zadania = machine.setdefault("zadania", [])
        zadania.append({"data": dlg.result, "typ_zadania": "przegląd", "uwagi": ""})


def _set_machine_image(parent, machine: dict):
    path = pick_machine_image(parent)
    if not path:
        return
    machine["image"] = os.path.normpath(path)


def _build_edit_footer(panel, machine: dict, on_changed):
    footer = ttk.Frame(panel)
    footer.grid(row=7, column=0, columnspan=2, sticky="we", padx=6, pady=(8, 0))
    footer.columnconfigure(2, weight=1)

    info_var = tk.StringVar()

    def _summary() -> str:
        tasks = machine.get("zadania") or []
        image_path = machine.get("image") or machine.get("obraz") or ""
        parts = [
            f"Przeglądy: {len(tasks)}" if tasks else "Przeglądy: brak",
            f"Zdjęcie: {os.path.basename(image_path)}" if image_path else "Zdjęcie: brak",
        ]
        return " | ".join(parts)

    def _trigger_changed() -> None:
        info_var.set(_summary())
        if callable(on_changed):
            on_changed()

    def _handle_add():
        before = list(machine.get("zadania") or [])
        _add_inspection_date(panel, machine)
        after = machine.get("zadania") or []
        if after != before:
            _trigger_changed()

    def _handle_image():
        before = machine.get("image") or machine.get("obraz")
        _set_machine_image(panel, machine)
        after = machine.get("image") or machine.get("obraz")
        if after != before:
            _trigger_changed()

    ttk.Button(
        footer,
        text="Dodaj przegląd (miesiąc/rok)",
        command=_handle_add,
    ).grid(row=0, column=0, sticky="w")
    ttk.Button(
        footer,
        text="Ustaw zdjęcie…",
        command=_handle_image,
    ).grid(row=0, column=1, sticky="w", padx=(8, 0))
    ttk.Label(footer, textvariable=info_var).grid(row=0, column=2, sticky="w", padx=(12, 0))

    _trigger_changed()
    return footer


def _save_machines(primary_path: str, rows: list[dict]) -> bool:
    try:
        payload = {"maszyny": rows}
        _safe_write_json(primary_path, payload)
        logger.info("[Maszyny] Zapisano %d rekordów -> %s", len(rows), primary_path)
        return True
    except Exception:
        logger.exception("[Maszyny] Błąd zapisu danych maszyn (%s)", primary_path)
        return False


_TREE_COLUMN_LAYOUT: Tuple[Tuple[str, str, int, str], ...] = (
    ("id", "ID", 110, "w"),
    ("nazwa", "Nazwa", 220, "w"),
    ("typ", "Typ", 120, "w"),
    ("status", "Status", 90, "center"),
    ("przeglad", "Najbliższy przegląd", 190, "center"),
    ("przeglad_status", "Status przeglądu", 150, "center"),
    ("dni", "Dni", 70, "center"),
)


def _ensure_tree_columns(tree: ttk.Treeview) -> None:
    columns = [cid for cid, _, _, _ in _TREE_COLUMN_LAYOUT]
    tree.configure(columns=columns)
    for cid, label, width, anchor in _TREE_COLUMN_LAYOUT:
        tree.heading(
            cid,
            text=label,
            command=lambda c=cid: _sort_treeview_by_column(tree, c),
        )
        tree.column(cid, width=width, anchor=anchor)


def _sort_treeview_by_column(tree: ttk.Treeview, column: str) -> None:
    """Sortuje tabelę maszyn po kliknięciu nagłówka kolumny."""

    columns = list(tree["columns"]) if tree["columns"] else []
    if column not in columns:
        return

    col_index = columns.index(column)
    reverse = bool(getattr(tree, "_wm_sort_reverse", False))
    last_column = getattr(tree, "_wm_sort_column", None)
    if last_column == column:
        reverse = not reverse
    else:
        reverse = False

    def _sort_key(item_id: str):
        values = tree.item(item_id, "values") or []
        value = values[col_index] if col_index < len(values) else ""
        text = str(value or "").strip()

        if column in ("id", "dni"):
            digits = "".join(ch for ch in text if ch.isdigit() or ch == "-")
            try:
                return (0, int(digits))
            except Exception:
                return (1, text.lower())

        return (0, text.lower())

    items = list(tree.get_children(""))
    items.sort(key=_sort_key, reverse=reverse)

    for index, item_id in enumerate(items):
        tree.move(item_id, "", index)

    setattr(tree, "_wm_sort_column", column)
    setattr(tree, "_wm_sort_reverse", reverse)


def _tree_insert_row(tree: ttk.Treeview, machine: dict) -> str:
    columns = list(tree["columns"]) if tree["columns"] else []
    summary = machine.get("__schedule_summary") if isinstance(machine, dict) else {}
    next_label = (summary or {}).get("next_label") or "—"
    status_label = (summary or {}).get("status_label") or "Brak danych"
    days = (summary or {}).get("days")

    def _value_for(col: str) -> str | int:
        if col == "dni":
            return days if days is not None else "—"
        if col == "przeglad":
            return next_label
        if col == "przeglad_status":
            return status_label
        if col == "id":
            return machine.get("id", "") or machine.get("nr_ewid", "") or ""
        if col == "status":
            return _machine_status_label(machine.get("status"))
        return machine.get(col, "")

    values = tuple(_value_for(col) for col in columns)
    identifier = str(machine.get("id") or machine.get("nr_ewid") or "")
    item_id = tree.insert("", "end", iid=identifier or None, values=values)
    tag_identifier = identifier or item_id
    machine_status_tag = _ensure_tree_machine_status_tag(tree, machine.get("status"))
    tags = [f"ROW::{tag_identifier}"]
    if machine_status_tag:
        tags.append(machine_status_tag)
    tree.item(item_id, tags=tuple(tags))
    return item_id


def _bind_tree_tooltips(tree: ttk.Treeview, rows: list[dict], root_widget: tk.Misc) -> None:
    tip = ImageTooltip(root_widget)
    id_by_item: dict[str, dict] = {}
    for item in tree.get_children(""):
        tags = tree.item(item, "tags") or ()
        tag = next((t for t in tags if t.startswith("ROW::")), None)
        if not tag:
            continue
        row_id = tag.split("::", 1)[1]
        machine = next(
            (m for m in rows if str(m.get("id") or m.get("nr_ewid") or "") == row_id),
            None,
        )
        if machine is None:
            values = tree.item(item, "values")
            if values:
                fallback_id = str(values[0])
                machine = next(
                    (
                        m
                        for m in rows
                        if str(m.get("id") or m.get("nr_ewid") or "") == fallback_id
                    ),
                    None,
                )
        if machine:
            id_by_item[item] = machine

    def _on_motion(event: tk.Event) -> None:
        item = tree.identify_row(event.y)
        if item and item in id_by_item:
            tip.show(event.x_root, event.y_root, id_by_item[item])
        else:
            tip.hide()

    def _on_leave(_event: tk.Event) -> None:
        tip.hide()

    tree.bind("<Motion>", _on_motion)
    tree.bind("<Leave>", _on_leave)


# ============================================================
# Renderer hali: kropki z etykietą w środku, drag&drop, tooltip z miniaturą
# ============================================================
class MachineHallRenderer:
    COLORS = {
        "SELECTED": "#60a5fa",
        "_": "#1f2937",
        "OK": "#16a34a",
        "WARN": "#ca8a04",
        "ALERT": "#dc2626",
    }
    RADIUS = 18

    def __init__(
        self,
        parent: tk.Misc,
        rows: List[Dict],
        cfg: dict | None = None,
        on_drag_commit=None,
        bg_path: str | None = None,
    ):
        self.parent = parent
        self.rows = rows or []
        self.cfg = cfg or {}
        self.on_drag_commit = on_drag_commit
        self.canvas = tk.Canvas(
            parent,
            bg=DEFAULT_BG_COLOR,
            highlightthickness=0,
            width=CANVAS_W,
            height=CANVAS_H,
        )
        self.nodes_by_id: dict[str, int] = {}
        self.text_by_id: dict[str, int] = {}
        self.labels_by_id: dict[str, int] = {}
        self.rows_by_id: dict[str, dict] = {}
        self.selected_id: Optional[str] = None
        self._drag_active = False
        self._drag_id: Optional[str] = None
        self._offset: tuple[int, int] = (0, 0)
        self._bg_path = bg_path
        self._bg_image_path: Optional[str] = None
        self._bg_img_pil: Optional["Image.Image"] = None
        self._bg_img_tk: Optional["ImageTk.PhotoImage"] = None
        self._bg_fallback: Optional[tk.PhotoImage] = None
        self._bg_w = 0
        self._bg_h = 0
        self._bg_anchor_xy: Tuple[int, int] = (0, 0)
        self._scale_x = 1.0
        self._scale_y = 1.0
        self.tip = ImageTooltip(parent)
        self.items_meta: dict[int, dict] = {}
        self._current_radius = self.RADIUS

    # ---------- public ----------
    def render(self) -> None:
        self.canvas.config(width=CANVAS_W, height=CANVAS_H)
        self.canvas.pack(fill="none", expand=False, padx=8, pady=8)
        self._load_background()
        self._draw_all()
        self._bind_drag()
        self.canvas.bind("<Motion>", self._on_canvas_motion, add="+")
        self.canvas.bind("<Leave>", lambda _e: self.tip.hide(), add="+")

    def select(self, machine_id: str | None):
        self.selected_id = machine_id
        self._redraw_selection()

    def update_rows(self, rows: List[Dict]):
        self.rows = rows or []
        self._draw_all()

    # ---------- internals ----------
    def _load_background(self):
        self._reset_background_state()

        path = None
        machines_cfg = self.cfg.get("machines") if isinstance(self.cfg, dict) else None
        if isinstance(machines_cfg, dict):
            candidate = machines_cfg.get("background_image")
            if isinstance(candidate, str) and candidate.strip():
                path = candidate
        if not path:
            cfg_paths = (self.cfg.get("paths", {}) or {})
            cfg_bg = (cfg_paths.get("hall", {}) or {}).get("background_image") or cfg_paths.get(
                "hall.background_image"
            )
            if isinstance(cfg_bg, str) and cfg_bg.strip():
                path = cfg_bg
        if not path and isinstance(self._bg_path, str):
            path = self._bg_path
        if not path:
            return
        if not os.path.isabs(path):
            cfg_context = self.cfg if isinstance(self.cfg, dict) else {}
            root = _coalesce_data_root(cfg_context)
            path = os.path.join(root, path)
        path = os.path.normpath(path)
        if not os.path.exists(path):
            logger.info("[Maszyny][HALL] Tło nie istnieje: %s", path)
            return

        self._bg_image_path = path
        self._load_bg_image_assets(path)

    def _reset_background_state(self) -> None:
        self._bg_image_path = None
        self._bg_img_pil = None
        self._bg_img_tk = None
        self._bg_fallback = None
        self._bg_w = 0
        self._bg_h = 0
        self._bg_anchor_xy = (0, 0)
        self._scale_x = 1.0
        self._scale_y = 1.0

    def _set_bg_geometry(self, width: int, height: int) -> None:
        self._bg_w = max(0, int(width))
        self._bg_h = max(0, int(height))
        off_x = max(0, (CANVAS_W - self._bg_w) // 2)
        off_y = max(0, (CANVAS_H - self._bg_h) // 2)
        self._bg_anchor_xy = (off_x, off_y)
        # Tło nie jest skalowane – współczynniki 1.0 pozwalają mapować px->canvas.
        self._scale_x = 1.0
        self._scale_y = 1.0

    def _load_bg_image_assets(self, path: str) -> None:
        if Image and ImageTk:
            try:
                img = Image.open(path)
            except Exception:
                img = None
            else:
                self._bg_img_pil = img
                self._set_bg_geometry(img.width, img.height)
                try:
                    self._bg_img_tk = ImageTk.PhotoImage(img)
                except Exception:
                    self._bg_img_tk = None
        if self._bg_img_tk is not None:
            self._bg_fallback = None
            return
        try:
            tk_img = tk.PhotoImage(file=path)
        except Exception:
            self._bg_fallback = None
            self._set_bg_geometry(0, 0)
            return
        self._bg_fallback = tk_img
        try:
            width = int(tk_img.width())
            height = int(tk_img.height())
        except Exception:
            width = height = 0
        self._set_bg_geometry(width, height)

    def _draw_background_and_grid(self) -> None:
        self.canvas.config(width=CANVAS_W, height=CANVAS_H)
        self.canvas.create_rectangle(
            0,
            0,
            CANVAS_W,
            CANVAS_H,
            fill=DEFAULT_BG_COLOR,
            outline="",
        )

        ax, ay = self._bg_anchor_xy
        if self._bg_img_tk is not None:
            self.canvas.create_image(ax, ay, image=self._bg_img_tk, anchor="nw")
        elif self._bg_fallback is not None:
            self.canvas.create_image(ax, ay, image=self._bg_fallback, anchor="nw")

        if self._bg_w > 0 and self._bg_h > 0:
            step_x = max(1, int(GRID_BASE_BG_PX_X * self._scale_x))
            step_y = max(1, int(GRID_BASE_BG_PX_Y * self._scale_y))

            x = ax
            while x <= ax + self._bg_w:
                self.canvas.create_line(x, ay, x, ay + self._bg_h, fill="#2a2a2a")
                x += step_x

            y = ay
            while y <= ay + self._bg_h:
                self.canvas.create_line(ax, y, ax + self._bg_w, y, fill="#2a2a2a")
                y += step_y

            self.canvas.create_rectangle(
                ax,
                ay,
                ax + self._bg_w,
                ay + self._bg_h,
                outline="#3a3a3a",
            )

    def _map_bg_to_canvas(self, x_bg: int, y_bg: int) -> Tuple[int, int]:
        ax, ay = self._bg_anchor_xy
        cx = int(ax + x_bg * self._scale_x)
        cy = int(ay + y_bg * self._scale_y)
        return self._clamp_to_canvas(cx, cy)

    def _map_canvas_to_bg(self, x_canvas: int, y_canvas: int) -> Tuple[int, int]:
        if self._bg_w > 0 and self._bg_h > 0:
            ax, ay = self._bg_anchor_xy
            try:
                bx = int(round((x_canvas - ax) / self._scale_x))
                by = int(round((y_canvas - ay) / self._scale_y))
            except ZeroDivisionError:
                bx, by = x_canvas, y_canvas
            bx = _safe_clamp(bx, 0, max(0, self._bg_w - 1))
            by = _safe_clamp(by, 0, max(0, self._bg_h - 1))
            return bx, by
        return x_canvas, y_canvas

    def _clamp_to_canvas(self, x: int, y: int) -> Tuple[int, int]:
        clamped_x = _safe_clamp(int(x), 0, CANVAS_W - 1)
        clamped_y = _safe_clamp(int(y), 0, CANVAS_H - 1)
        return clamped_x, clamped_y

    def _status_color(self, status):
        """Kolor kółka maszyny zgodny z polskimi statusami WM."""

        return _machine_status_color(status)

    def _node_center(self, r: Dict, idx: int, radius: int) -> tuple[int, int]:
        cols, cell_w, cell_h, pad_x, pad_y = 6, 120, 110, 70, 70
        x, y = r.get("x"), r.get("y")
        if isinstance(x, int) and isinstance(y, int):
            margin = max(radius, 4)
            if self._bg_w > 0 and self._bg_h > 0:
                cx, cy = self._map_bg_to_canvas(x, y)
                ax, ay = self._bg_anchor_xy
                min_x = ax + margin
                max_x = ax + self._bg_w - margin
                min_y = ay + margin
                max_y = ay + self._bg_h - margin
                if max_x < min_x:
                    min_x, max_x = margin, max(margin, CANVAS_W - margin)
                if max_y < min_y:
                    min_y, max_y = margin, max(margin, CANVAS_H - margin)
            else:
                width, height = _canvas_bounds(self.canvas)
                min_x = margin
                min_y = margin
                max_x = max(margin, width - margin)
                max_y = max(margin, height - margin)
                cx, cy = x, y
            return (
                _safe_clamp(int(cx), int(min_x), int(max_x)),
                _safe_clamp(int(cy), int(min_y), int(max_y)),
            )
        gx, gy = idx % cols, idx // cols
        cx, cy = pad_x + gx * cell_w, pad_y + gy * cell_h
        width, height = _canvas_bounds(self.canvas)
        margin = max(radius, 4)
        max_x = max(margin, width - margin)
        max_y = max(margin, height - margin)
        return (
            _safe_clamp(int(cx), margin, max_x),
            _safe_clamp(int(cy), margin, max_y),
        )

    def _short_id(self, mid: str) -> str:
        mid = (mid or "").strip()
        if len(mid) <= 5:
            return mid
        # skróć typu ABC-1234 -> A-1234
        parts = mid.split("-")
        if len(parts) >= 2 and parts[0]:
            return f"{parts[0][0]}-{parts[-1][:4]}"
        return mid[:5]

    def _draw_all(self):
        self.canvas.delete("all")
        self.nodes_by_id.clear()
        self.text_by_id.clear()
        self.labels_by_id.clear()
        self.rows_by_id.clear()
        self.items_meta.clear()

        radius = self._resolve_radius()
        self._current_radius = radius

        self._draw_background_and_grid()

        label_mode = self._label_mode()
        for idx, r in enumerate(self.rows):
            if not isinstance(r, dict):
                continue
            mid = str(r.get("id") or r.get("nr_ewid") or f"row{idx}")
            self.rows_by_id[mid] = r

            cx, cy = self._node_center(r, idx, radius)
            node = self.canvas.create_oval(
                cx - radius,
                cy - radius,
                cx + radius,
                cy + radius,
                fill=self._status_color(r.get("status")),
                outline="#0b0c0f",
                width=1,
            )
            self.nodes_by_id[mid] = node
            self.items_meta[node] = r

            label_text = _map_label_text(r, label_mode).strip()
            label_text = label_text[:6]
            font_size = max(8, min(14, int(radius * 0.55)))
            text_id = self.canvas.create_text(
                cx,
                cy,
                text=label_text,
                fill="#0b0c0f",
                font=("TkDefaultFont", font_size, "bold"),
            )
            self.text_by_id[mid] = text_id
            self.items_meta[text_id] = r

            label_id = _render_days_label_on_canvas(
                self.canvas, cx, cy + radius + 14, r
            )
            if label_id is not None:
                self.labels_by_id[mid] = label_id
                self.items_meta[label_id] = r

        self._redraw_selection()

    def _redraw_selection(self):
        for mid, node in self.nodes_by_id.items():
            sel = mid == self.selected_id
            try:
                self.canvas.itemconfigure(
                    node,
                    outline=self.COLORS["SELECTED"] if sel else "#0b0c0f",
                    width=3 if sel else 1,
                )
            except Exception:
                pass

    def _find_node_at(self, x: int, y: int) -> Optional[str]:
        radius = self._current_radius or self.RADIUS
        items = self.canvas.find_overlapping(
            x - radius,
            y - radius,
            x + radius,
            y + radius,
        )
        inv = {v: k for k, v in self.nodes_by_id.items()}
        for it in items:
            if it in inv:
                return inv[it]
        return None

    def _bind_drag(self):
        self.canvas.bind("<Button-1>", self._on_press, add="+")
        self.canvas.bind("<B1-Motion>", self._on_motion, add="+")
        self.canvas.bind("<ButtonRelease-1>", self._on_release, add="+")

    def _on_press(self, ev):
        mid = self._find_node_at(ev.x, ev.y)
        self._drag_active = bool(mid)
        self._drag_id = mid
        if mid:
            self.select(mid)
            node = self.nodes_by_id.get(mid)
            if node:
                x1, y1, x2, y2 = self.canvas.coords(node)
                cx = (x1 + x2) / 2
                cy = (y1 + y2) / 2
                self._offset = (int(cx - ev.x), int(cy - ev.y))
        else:
            self._offset = (0, 0)

    def _move_group(self, mid: str, cx: int, cy: int):
        node = self.nodes_by_id.get(mid)
        radius = self._current_radius or self.RADIUS
        margin = max(radius, 4)
        if self._bg_w > 0 and self._bg_h > 0:
            ax, ay = self._bg_anchor_xy
            min_x = ax + margin
            min_y = ay + margin
            max_x = ax + self._bg_w - margin
            max_y = ay + self._bg_h - margin
            if max_x < min_x:
                width, height = _canvas_bounds(self.canvas)
                min_x = margin
                max_x = max(margin, width - margin)
            if max_y < min_y:
                width, height = _canvas_bounds(self.canvas)
                min_y = margin
                max_y = max(margin, height - margin)
        else:
            width, height = _canvas_bounds(self.canvas)
            min_x = margin
            min_y = margin
            max_x = max(margin, width - margin)
            max_y = max(margin, height - margin)
        cx = _safe_clamp(cx, int(min_x), int(max_x))
        cy = _safe_clamp(cy, int(min_y), int(max_y))
        if node:
            self.canvas.coords(
                node,
                cx - radius,
                cy - radius,
                cx + radius,
                cy + radius,
            )
        t = self.text_by_id.get(mid)
        if t:
            self.canvas.coords(t, cx, cy)
        lab = self.labels_by_id.get(mid)
        if lab:
            self.canvas.coords(lab, cx, cy + (radius + 14))

    def _on_motion(self, ev):
        if not self._drag_active or not self._drag_id:
            return
        self.tip.hide()
        offx, offy = self._offset
        cx, cy = ev.x + offx, ev.y + offy
        grid = 10
        cx = int(round(cx / grid) * grid)
        cy = int(round(cy / grid) * grid)
        self._move_group(self._drag_id, cx, cy)

    def _on_release(self, ev):
        if not self._drag_active or not self._drag_id:
            return
        node = self.nodes_by_id.get(self._drag_id)
        if not node:
            self._drag_active = False
            self._drag_id = None
            return
        x1, y1, x2, y2 = self.canvas.coords(node)
        cx, cy = int((x1 + x2) / 2), int((y1 + y2) / 2)
        bx, by = self._map_canvas_to_bg(cx, cy)

        # aktualizuj model i zapisz
        r = self.rows_by_id.get(self._drag_id)
        if r is not None:
            r["x"], r["y"] = bx, by
        if callable(self.on_drag_commit):
            try:
                self.on_drag_commit(self._drag_id, bx, by)
            except Exception:
                logger.exception("[Maszyny][HALL] Błąd zapisu po drag&drop")

        self._drag_active = False
        self._drag_id = None

    def _on_canvas_motion(self, event: tk.Event) -> None:
        if self._drag_active:
            return
        current = self.canvas.find_withtag("current")
        if not current:
            self.tip.hide()
            return
        item_id = current[0]
        machine = self.items_meta.get(item_id)
        if machine is None:
            pair = self._pair(item_id)
            if pair is not None:
                machine = self.items_meta.get(pair)
        if machine:
            self.tip.show(event.x_root, event.y_root, machine)
        else:
            self.tip.hide()

    def _pair(self, item_id: int) -> Optional[int]:
        bbox = self.canvas.bbox(item_id)
        if not bbox:
            return None
        x1, y1, x2, y2 = bbox
        nearby = self.canvas.find_overlapping(x1 - 2, y1 - 2, x2 + 2, y2 + 2)
        for candidate in nearby:
            if candidate != item_id and candidate in self.items_meta:
                return candidate
        return None

    def _label_mode(self) -> str:
        machines_cfg = self.cfg.get("machines") if isinstance(self.cfg, dict) else {}
        mode = "id"
        if isinstance(machines_cfg, dict):
            raw = (machines_cfg.get("map_label") or "id").strip().lower()
            if raw in {"id", "typ", "nazwa"}:
                mode = raw
        return mode

    def _resolve_radius(self) -> int:
        machines_cfg = self.cfg.get("machines") if isinstance(self.cfg, dict) else {}
        if isinstance(machines_cfg, dict):
            try:
                radius = int(machines_cfg.get("map_dot_radius") or self.RADIUS)
            except Exception:
                radius = self.RADIUS
        else:
            radius = self.RADIUS
        radius = radius or self.RADIUS
        return max(10, min(60, radius))

# ============================================================
# Reszta panelu — zostaje jak w Twojej wersji R-06Jc,
# poniżej fragmenty, które muszą zapewnić domyślne nr_hali=1
# przy edycji/zapisie oraz integrację z rendererem.
# ============================================================
def _build_tree(parent: tk.Misc, rows: List[Dict]) -> ttk.Treeview:
    tree = ttk.Treeview(
        parent,
        columns=tuple(cid for cid, _, _, _ in _TREE_COLUMN_LAYOUT),
        show="headings",
        height=18,
    )
    _ensure_tree_columns(tree)
    try:
        style = ttk.Style(tree)
        style.configure("Maszyny.Treeview", font=("Segoe UI", 11), rowheight=30)
        style.configure("Maszyny.Treeview.Heading", font=("Segoe UI", 11, "bold"))
        tree.configure(style="Maszyny.Treeview")
    except Exception:
        pass
    for r in rows:
        _tree_insert_row(tree, r)
    tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    return tree


def _detect_real_source(rows_from_fallback: List[Dict], primary_path: str, cfg: dict) -> str:
    primary_rows, _ = load_machines(primary_path)
    if rows_from_fallback and not primary_rows:
        legacy_path = resolve_rel(cfg, r"maszyny.json")
        legacy_rows, _ = load_machines(legacy_path)
        if legacy_rows:
            return legacy_path
    return primary_path


def _open_machines_panel(
    root: tk.Misc,
    container: tk.Misc,
    Renderer=None,
    *,
    initial_machine_id: str = "",
):
    for child in container.winfo_children():
        child.destroy()

    paned = ttk.Panedwindow(container, orient="horizontal")
    paned.pack(fill="both", expand=True)
    left, right = ttk.Frame(paned), ttk.Frame(paned)
    paned.add(left, weight=1)
    paned.add(right, weight=1)

    toolbar = ttk.Frame(left)
    toolbar.pack(fill="x", padx=8, pady=(8, 0))
    info = tk.StringVar(value="Maszyny")
    ttk.Label(toolbar, textvariable=info).pack(side="left")

    search_var = tk.StringVar(value="")
    ttk.Label(toolbar, text="Szukaj:").pack(side="left", padx=(12, 4))
    entry_search = ttk.Entry(toolbar, textvariable=search_var, width=28)
    entry_search.pack(side="left", padx=(0, 6))
    btn_clear_search = ttk.Button(toolbar, text="Wyczyść")
    btn_clear_search.pack(side="left", padx=(0, 8))

    filter_var = tk.StringVar(value="Wszystkie")
    machine_mode_var = tk.StringVar(value="Użytkowanie")

    def _is_machine_edit_mode() -> bool:
        return machine_mode_var.get() == "Edycja maszyn"

    def _require_machine_edit_mode() -> bool:
        if _is_machine_edit_mode():
            return True
        messagebox.showinfo(
            "Maszyny",
            (
                "Jesteś w trybie Użytkowanie.\n\n"
                "Przełącz na „Edycja maszyn”, żeby dodawać, edytować, usuwać "
                "albo importować dane techniczne maszyn."
            ),
            parent=root,
        )
        return False

    ttk.Label(toolbar, text="Filtr:").pack(side="left", padx=(4, 4))
    filter_box = ttk.Combobox(
        toolbar,
        state="readonly",
        width=14,
        values=("Wszystkie", "Po terminie", "Wkrótce", "Planowane", "Wykonane"),
        textvariable=filter_var,
    )
    filter_box.pack(side="left")

    ttk.Label(toolbar, text="Tryb:").pack(side="left", padx=(8, 4))
    mode_box = ttk.Combobox(
        toolbar,
        textvariable=machine_mode_var,
        values=("Użytkowanie", "Edycja maszyn"),
        state="readonly",
        width=18,
    )
    mode_box.pack(side="left", padx=(0, 12))

    mode_hint_var = tk.StringVar(
        value="Tryb bezpieczny: można przeglądać i obsługiwać statusy."
    )
    ttk.Label(toolbar, textvariable=mode_hint_var).pack(side="left", padx=(0, 12))

    actions_toolbar = ttk.Frame(left)
    actions_toolbar.pack(fill="x", padx=8, pady=(4, 0))

    btn_import = ttk.Button(toolbar, text="Importuj harmonogram…")
    # Stary import harmonogramu ukryty. Docelowym modelem są machine["reviews"].

    btn_add, btn_edit, btn_del, btn_save = (
        ttk.Button(actions_toolbar, text=text)
        for text in ("Dodaj", "Edytuj", "Usuń", "Zapisz")
    )
    btn_change_status = ttk.Button(actions_toolbar, text="Zmień status")
    edit_mode_buttons = [btn_add, btn_edit, btn_del, btn_save]
    for button in (btn_add, btn_change_status, btn_edit, btn_del, btn_save):
        button.pack(side="left", padx=(0, 6))

    def _refresh_machine_mode_ui(*_args) -> None:
        edit_mode = _is_machine_edit_mode()
        for button in edit_mode_buttons:
            button.state(["!disabled"] if edit_mode else ["disabled"])
        mode_hint_var.set(
            "Tryb edycji: można zmieniać dane techniczne maszyn."
            if edit_mode
            else "Tryb bezpieczny: można przeglądać i obsługiwać statusy."
        )

    schedule_info = tk.StringVar(value="")
    ttk.Label(left, textvariable=schedule_info).pack(fill="x", padx=8, pady=(4, 4))

    cfg: Dict[str, Any] = {}
    try:
        cfg = get_config()
    except Exception:
        logger.exception("[Maszyny] Nie udało się wczytać konfiguracji.")

    cfg_manager = None
    try:
        from config_manager import ConfigManager

        cfg_manager = ConfigManager()
    except Exception:
        cfg_manager = None

    rows, primary_path = load_machines_rows_with_fallback(cfg, resolve_rel)
    try:
        from gui_panel import wm_set_module_source
        wm_set_module_source(root, "Maszyny", primary_path)
    except Exception:
        pass

    def _cards_output_dir() -> Path:
        try:
            from config_manager import ConfigManager

            base = Path(ConfigManager().path_root("wydruki", "karty"))
        except Exception:
            root = str(os.environ.get("WM_ROOT") or "").strip()
            base = (Path(root) if root else Path.cwd()) / "wydruki" / "karty"
        base.mkdir(parents=True, exist_ok=True)
        return base

    def _print_blank_machine_card_from_machines() -> None:
        try:
            from machine_card_pdf import generate_blank_machine_card

            generate_blank_machine_card(
                _cards_output_dir(),
                open_after=True,
            )
        except Exception as exc:
            messagebox.showerror(
                "Maszyny",
                f"Nie udało się wygenerować pustej karty maszyny:\n{exc}",
                parent=root if hasattr(root, "winfo_exists") else None,
            )

    had_rows = bool(rows)
    rows = ensure_machines_sample_if_empty(rows, primary_path)
    source_path = _detect_real_source(rows, primary_path, cfg)
    rows_cache: List[Dict] = list(rows)

    schedule_year = SCHEDULE_YEAR
    schedule_path = resolve_schedule_path(schedule_year, cfg_manager)
    schedule_payload = _safe_read_json(schedule_path, default={})
    schedule_meta: Dict[str, Any] = {}
    raw_schedule_entries: List[Dict[str, Any]] = []
    if isinstance(schedule_payload, dict):
        schedule_meta = {k: v for k, v in schedule_payload.items() if k != "entries"}
        raw_schedule_entries = [
            entry for entry in schedule_payload.get("entries", []) if isinstance(entry, dict)
        ]
        if schedule_meta.get("year"):
            try:
                schedule_year = int(schedule_meta["year"])
            except Exception:
                schedule_year = SCHEDULE_YEAR
    elif isinstance(schedule_payload, list):
        raw_schedule_entries = [entry for entry in schedule_payload if isinstance(entry, dict)]
        schedule_meta = {"year": schedule_year}
    else:
        schedule_meta = {"year": schedule_year}
    if schedule_year != SCHEDULE_YEAR:
        schedule_path = resolve_schedule_path(schedule_year, cfg_manager)

    schedule_entries: List[Dict[str, Any]] = []
    for raw_entry in raw_schedule_entries:
        entry = _normalize_schedule_entry(raw_entry)
        if not entry:
            continue
        if raw_entry.get("card"):
            entry["card"] = str(raw_entry.get("card")).strip()
        if raw_entry.get("completed_at") and "completed_at" not in entry:
            entry["completed_at"] = str(raw_entry.get("completed_at"))
        schedule_entries.append(entry)

    schedule_meta.setdefault("year", schedule_year)
    if schedule_meta.get("source") is None:
        schedule_meta["source"] = (
            os.path.basename(schedule_path) if os.path.exists(schedule_path) else ""
        )
    schedule_meta.setdefault("imported_at", schedule_meta.get("imported_at"))
    schedule_meta.setdefault("updated_at", schedule_meta.get("updated_at"))

    _attach_schedule(rows_cache, schedule_entries)
    for row in rows_cache:
        if isinstance(row, dict):
            row["__schedule_summary"] = _combined_machine_schedule_summary(row)
    visible_rows: List[Dict] = list(rows_cache)

    info.set(
        f"Wczytano {len(rows_cache)} maszyn." if had_rows else "Brak danych – dodano przykładowe pozycje."
    )

    machine_actions = ttk.Frame(left)
    machine_actions.pack(fill="x", pady=(0, 6))
    ttk.Button(
        machine_actions,
        text="Drukuj pustą kartę maszyny",
        command=_print_blank_machine_card_from_machines,
    ).pack(side="left")

    tree = _build_tree(left, visible_rows)
    _bind_tree_tooltips(tree, visible_rows, root)

    selected_machine_id: Optional[str] = None
    hall: MachineHallRenderer | None = None
    upcoming_items: Dict[str, Dict[str, Any]] = {}
    history_items: Dict[str, Dict[str, Any]] = {}

    def _refresh_schedule_info() -> None:
        overdue = sum(1 for row in rows_cache if _schedule_status_key(row) == "overdue")
        soon = sum(1 for row in rows_cache if _schedule_status_key(row) == "soon")
        planned = sum(1 for row in rows_cache if _schedule_status_key(row) == "ok")
        schedule_info.set(
            f"Przeglądy maszyn: {overdue} po terminie • "
            f"{soon} w ciągu {SCHEDULE_SOON_THRESHOLD_DAYS} dni • "
            f"{planned} później"
        )

    def _show_review_notice_once() -> None:
        if initial_machine_id:
            return
        overdue = sum(1 for row in rows_cache if _schedule_status_key(row) == "overdue")
        soon = sum(1 for row in rows_cache if _schedule_status_key(row) == "soon")
        if overdue <= 0 and soon <= 0:
            return
        try:
            target = root.winfo_toplevel()
        except Exception:
            target = root
        notice_key = f"{dt.date.today().isoformat()}:{overdue}:{soon}"
        if getattr(target, "_wm_machine_review_notice_key", "") == notice_key:
            return
        try:
            setattr(target, "_wm_machine_review_notice_key", notice_key)
        except Exception:
            pass
        lines = []
        if overdue:
            lines.append(f"Po terminie: {overdue}")
        if soon:
            lines.append(
                f"W ciągu {SCHEDULE_SOON_THRESHOLD_DAYS} dni: {soon}"
            )
        show_now = messagebox.askyesno(
            "Przeglądy maszyn",
            "\n".join(lines)
            + "\n\nOdpowiednie Dyspozycje są tworzone automatycznie. "
            + "Pokazać te maszyny teraz?",
            parent=target if getattr(target, "tk", None) is not None else None,
        )
        if show_now:
            filter_var.set("Po terminie" if overdue else "Wkrótce")
            _apply_filter()

    def _update_info() -> None:
        info.set(f"Wczytano {len(rows_cache)} maszyn • widocznych: {len(visible_rows)}")

    def _recompute_visible_rows() -> None:
        nonlocal visible_rows
        mode = filter_var.get()
        query = search_var.get().strip().lower()

        def predicate(machine: Dict[str, Any]) -> bool:
            key = _schedule_status_key(machine)
            if mode == "Po terminie":
                return key == "overdue"
            if mode == "Wkrótce":
                return key == "soon"
            if mode == "Planowane":
                return key == "ok"
            if mode == "Wykonane":
                return key == "done"
            return True

        def matches_search(machine: Dict[str, Any]) -> bool:
            if not query:
                return True
            summary = machine.get("__schedule_summary") or {}
            haystack = " ".join(
                str(value or "")
                for value in (
                    machine.get("id"),
                    machine.get("nr_ewid"),
                    machine.get("nr"),
                    machine.get("nazwa"),
                    machine.get("typ"),
                    machine.get("status"),
                    machine.get("lokalizacja"),
                    summary.get("next_label"),
                    summary.get("status_label"),
                )
            ).lower()
            return query in haystack

        visible_rows = [
            row for row in rows_cache if predicate(row) and matches_search(row)
        ]

    def _focus_first_machine_row() -> bool:
        children = tree.get_children("")
        if not children:
            return False
        first = children[0]
        tree.selection_set(first)
        tree.focus(first)
        tree.see(first)
        return True

    def _find_machine(machine_id: Optional[str]) -> Optional[Dict]:
        if not machine_id:
            return None
        return next(
            (
                row
                for row in rows_cache
                if str(row.get("id") or row.get("nr_ewid") or "") == machine_id
            ),
            None,
        )

    def _save_schedule_state() -> None:
        nonlocal schedule_path, schedule_year
        try:
            schedule_year = int(schedule_meta.get("year", schedule_year) or schedule_year)
        except Exception:
            schedule_year = SCHEDULE_YEAR
        schedule_path = resolve_schedule_path(schedule_year, cfg_manager)
        payload = dict(schedule_meta)
        payload["year"] = schedule_year
        payload["entries"] = _serialize_schedule_entries(schedule_entries)
        _safe_write_json(schedule_path, payload)

    def _refresh_tree() -> None:
        tree.delete(*tree.get_children())
        _ensure_tree_columns(tree)
        for row in visible_rows:
            _tree_insert_row(tree, row)
        _bind_tree_tooltips(tree, visible_rows, root)
        visible_ids = {
            str(row.get("id") or row.get("nr_ewid") or "")
            for row in visible_rows
        }
        if selected_machine_id and selected_machine_id in visible_ids:
            try:
                tree.selection_set(selected_machine_id)
            except Exception:
                tree.selection_remove(tree.selection())
        else:
            tree.selection_remove(tree.selection())
        _update_info()

    def _reload_from(path: str) -> List[Dict]:
        payload = _safe_read_json(path, default=[])
        return normalize_doc_list_or_dict(payload, "maszyny", fallback_keys=("machines",))

    def _on_rows_changed() -> None:
        _attach_schedule(rows_cache, schedule_entries)
        for row in rows_cache:
            if isinstance(row, dict):
                row["__schedule_summary"] = _combined_machine_schedule_summary(row)
        _recompute_visible_rows()
        _refresh_tree()
        _refresh_schedule_info()
        if hall is not None:
            hall.update_rows(rows_cache)
        if selected_machine_id:
            machine = _find_machine(selected_machine_id)
            _populate_details(machine)
        else:
            _populate_details(None)

    def _on_schedule_changed(save: bool = True) -> None:
        schedule_meta.setdefault("year", schedule_year)
        schedule_meta["updated_at"] = dt.datetime.now().isoformat()
        _on_rows_changed()
        _refresh_schedule_info()
        if save:
            _save_schedule_state()

    def _save_rows(rows_to_save: List[Dict]) -> List[Dict]:
        nonlocal source_path
        cleaned = [_strip_schedule_fields(row) for row in rows_to_save]
        if os.path.normpath(source_path) != os.path.normpath(primary_path):
            legacy_rows = _reload_from(source_path)
            prim_rows = _reload_from(primary_path)
            merged = merge_rows_union_by_id(prim_rows, legacy_rows)
            merged = merge_rows_union_by_id(merged, cleaned)
            if _save_machines(primary_path, merged):
                source_path = primary_path
                _attach_schedule(merged, schedule_entries)
                return merged
            return rows_to_save
        if _save_machines(primary_path, cleaned):
            _attach_schedule(cleaned, schedule_entries)
            return cleaned
        return rows_to_save

    def _apply_filter(*_args) -> None:
        _recompute_visible_rows()
        _refresh_tree()
        if selected_machine_id and not any(
            str(row.get("id") or row.get("nr_ewid") or "") == selected_machine_id
            for row in visible_rows
        ):
            _set_selected_machine(None)
        else:
            if selected_machine_id:
                _populate_details(_find_machine(selected_machine_id))
            else:
                _populate_details(None)

    def _apply_search(*_args) -> None:
        _apply_filter()
        if not selected_machine_id:
            _focus_first_machine_row()

    def _clear_search() -> None:
        search_var.set("")
        _apply_filter()
        _focus_first_machine_row()

    def _on_search_enter(_event=None) -> str:
        children = tree.get_children("")
        if not children:
            return "break"
        current = tree.selection()
        if not current or current[0] != children[0]:
            _focus_first_machine_row()
        else:
            _on_edit()
        return "break"

    def _drag_commit(mid: str, x: int, y: int) -> None:
        nonlocal rows_cache
        update = None
        for row in rows_cache:
            if str(row.get("id") or row.get("nr_ewid") or "") == mid:
                update = dict(row)
                update["x"], update["y"] = x, y
                if "nr_hali" not in update or update.get("nr_hali") in (None, ""):
                    update["nr_hali"] = "1"
                break
        if update is None:
            return
        new_rows = upsert_machine(rows_cache, update)
        persisted = _save_rows(new_rows)
        rows_cache = list(persisted)
        _on_rows_changed()

    hall = MachineHallRenderer(right, rows_cache, cfg=cfg, on_drag_commit=_drag_commit)
    hall.render()

    details = ttk.LabelFrame(right, text="Przeglądy")
    # Stary panel harmonogramu 2025 jest wyłączony, żeby nie mieszać go z nowym
    # modelem machine["reviews"] w oknie Użytkowanie maszyny.
    # details.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    summary_var = tk.StringVar(value="Wybierz maszynę, aby zobaczyć harmonogram.")
    ttk.Label(details, textvariable=summary_var).pack(fill="x", padx=8, pady=(6, 4))

    columns_details = ("data", "typ", "status", "uwagi")
    column_setup = {
        "data": ("Data", 110, "center"),
        "typ": ("Typ", 120, "center"),
        "status": ("Status", 160, "w"),
        "uwagi": ("Uwagi", 220, "w"),
    }

    upcoming_section = ttk.LabelFrame(details, text="Zaplanowane")
    upcoming_section.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    upcoming_tree = ttk.Treeview(upcoming_section, columns=columns_details, show="headings", height=6)
    for cid, (label, width, anchor) in column_setup.items():
        upcoming_tree.heading(cid, text=label)
        upcoming_tree.column(cid, width=width, anchor=anchor)
    upcoming_tree.pack(fill="both", expand=True, padx=6, pady=(0, 4))

    upcoming_buttons = ttk.Frame(upcoming_section)
    upcoming_buttons.pack(fill="x", padx=6, pady=(0, 4))
    btn_mark_done = ttk.Button(upcoming_buttons, text="Oznacz jako wykonany", state="disabled")
    btn_mark_done.pack(side="left")
    btn_assign_card = ttk.Button(upcoming_buttons, text="Przypisz kartę…", state="disabled")
    btn_assign_card.pack(side="left", padx=(6, 0))
    btn_open_card = ttk.Button(upcoming_buttons, text="Otwórz kartę", state="disabled")
    btn_open_card.pack(side="left", padx=(6, 0))

    history_section = ttk.LabelFrame(details, text="Historia")
    history_section.pack(fill="both", expand=True, padx=8, pady=(0, 8))
    history_tree = ttk.Treeview(history_section, columns=columns_details, show="headings", height=5)
    for cid, (label, width, anchor) in column_setup.items():
        history_tree.heading(cid, text=label)
        history_tree.column(cid, width=width, anchor=anchor)
    history_tree.pack(fill="both", expand=True, padx=6, pady=(0, 4))

    history_buttons = ttk.Frame(history_section)
    history_buttons.pack(fill="x", padx=6, pady=(0, 4))
    btn_restore = ttk.Button(history_buttons, text="Przywróć jako planowany", state="disabled")
    btn_restore.pack(side="left")

    def _selected_upcoming_entry() -> Optional[Dict[str, Any]]:
        sel = upcoming_tree.selection()
        if not sel:
            return None
        return upcoming_items.get(sel[0])

    def _selected_history_entry() -> Optional[Dict[str, Any]]:
        sel = history_tree.selection()
        if not sel:
            return None
        return history_items.get(sel[0])

    def _selected_schedule_entry() -> Optional[Dict[str, Any]]:
        entry = _selected_upcoming_entry()
        if entry is not None:
            return entry
        return _selected_history_entry()

    def _update_detail_buttons() -> None:
        if _selected_upcoming_entry() is not None:
            btn_mark_done.state(["!disabled"])
        else:
            btn_mark_done.state(["disabled"])
        if _selected_history_entry() is not None:
            btn_restore.state(["!disabled"])
        else:
            btn_restore.state(["disabled"])
        entry = _selected_schedule_entry()
        if entry is not None:
            btn_assign_card.state(["!disabled"])
            if entry.get("card"):
                btn_open_card.state(["!disabled"])
            else:
                btn_open_card.state(["disabled"])
        else:
            btn_assign_card.state(["disabled"])
            btn_open_card.state(["disabled"])

    def _populate_details(machine: Optional[Dict]) -> None:
        upcoming_items.clear()
        history_items.clear()
        for tree_view in (upcoming_tree, history_tree):
            for item in tree_view.get_children():
                tree_view.delete(item)
        if not machine:
            summary_var.set("Wybierz maszynę, aby zobaczyć harmonogram.")
            _update_detail_buttons()
            return
        summary = machine.get("__schedule_summary") or {}
        summary_var.set(summary.get("status_text") or "Brak danych harmonogramu")
        for entry in summary.get("upcoming", []):
            date_text = entry.get("date") or "—"
            typ = entry.get("type") or ""
            status_text, _status_key = _describe_entry_status(entry)
            notes = entry.get("notes") or ""
            iid = upcoming_tree.insert("", "end", values=(date_text, typ, status_text, notes))
            upcoming_items[iid] = entry
        for entry in summary.get("history", []):
            date_text = entry.get("date") or "—"
            typ = entry.get("type") or ""
            status_text, _status_key = _describe_entry_status(entry)
            notes = entry.get("notes") or ""
            iid = history_tree.insert("", "end", values=(date_text, typ, status_text, notes))
            history_items[iid] = entry
        _update_detail_buttons()

    def _set_selected_machine(machine_id: Optional[str]) -> None:
        nonlocal selected_machine_id
        selected_machine_id = machine_id
        if hall is not None:
            hall.select(machine_id)
        _populate_details(_find_machine(machine_id))

    def _mark_done() -> None:
        entry = _selected_upcoming_entry()
        if not entry:
            return
        entry["status"] = "wykonany"
        entry["completed_at"] = dt.datetime.now().isoformat()
        _on_schedule_changed()

    def _restore_plan() -> None:
        entry = _selected_history_entry()
        if not entry:
            return
        entry["status"] = "planowany"
        entry.pop("completed_at", None)
        _on_schedule_changed()

    def _assign_card() -> None:
        entry = _selected_schedule_entry()
        if not entry:
            return
        path = filedialog.askopenfilename(
            parent=root,
            title="Wybierz kartę przeglądu",
            filetypes=(("Dokumenty", "*.pdf;*.doc;*.docx;*.xlsx;*.xls;*.txt"), ("Wszystkie pliki", "*.*")),
        )
        if not path:
            return
        stored = _resolve_card_storage(path, cfg_manager)
        entry["card"] = stored
        _on_schedule_changed()

    def _open_selected_card() -> None:
        entry = _selected_schedule_entry()
        if not entry or not entry.get("card"):
            messagebox.showinfo("Karta przeglądu", "Brak przypisanej karty do tego wpisu.")
            return
        absolute = _resolve_card_absolute(str(entry.get("card")), cfg_manager)
        if not absolute or not os.path.exists(absolute):
            messagebox.showerror("Karta przeglądu", f"Plik nie istnieje: {absolute}")
            return
        if not _open_external(absolute):
            messagebox.showerror("Karta przeglądu", "Nie udało się otworzyć pliku.")

    def _on_upcoming_select(_event=None) -> None:
        if upcoming_tree.selection():
            history_tree.selection_remove(history_tree.selection())
        _update_detail_buttons()

    def _on_history_select(_event=None) -> None:
        if history_tree.selection():
            upcoming_tree.selection_remove(upcoming_tree.selection())
        _update_detail_buttons()

    def _selected_id() -> Optional[str]:
        sel = tree.selection()
        return str(sel[0]) if sel else None

    def _do_import() -> None:
        nonlocal schedule_entries, schedule_year, schedule_path
        if not _require_machine_edit_mode():
            return
        path = filedialog.askopenfilename(
            parent=root,
            title="Wybierz plik harmonogramu",
            filetypes=(("Pliki Excel", "*.xlsx;*.xls;*.xlsm"), ("Wszystkie pliki", "*.*")),
        )
        if not path:
            return
        try:
            new_entries, meta = _import_schedule_from_excel(path)
        except RuntimeError as exc:
            messagebox.showerror("Import harmonogramu", str(exc))
            return
        except ValueError as exc:
            messagebox.showerror("Import harmonogramu", str(exc))
            return
        merged = _merge_schedule_status(new_entries, schedule_entries)
        schedule_entries = list(merged)
        schedule_meta["year"] = meta.get("year", schedule_meta.get("year", schedule_year))
        schedule_year = int(schedule_meta.get("year", schedule_year) or schedule_year)
        schedule_meta["source"] = meta.get("source", os.path.basename(path))
        schedule_meta["imported_at"] = meta.get("imported_at")
        schedule_path = resolve_schedule_path(schedule_year, cfg_manager)
        _on_schedule_changed()
        messagebox.showinfo("Import harmonogramu", f"Zaimportowano {len(schedule_entries)} wpisów.")

    class MachineEditDialog(tk.Toplevel):
        STATUSES = MACHINE_STATUS_EDIT_VALUES

        def __init__(self, master: tk.Misc, row: dict | None, on_ok):
            super().__init__(master)
            self.title("Edycja maszyny")
            self.geometry("1100x720")
            self.minsize(980, 620)
            self.resizable(True, True)
            self.transient(master)
            self.grab_set()
            self._row = dict(row or {})
            self._on_ok = on_ok
            self._actor = _active_login_for_machine(master)
            self._old_status = _normalize_machine_status(self._row.get("status"))
            self._dirty = False

            def _mark_dirty(*_args) -> None:
                self._dirty = True

            def _confirm_close() -> None:
                if self._dirty:
                    if not messagebox.askyesno(
                        "Edycja maszyny",
                        "Masz niezapisane zmiany. Zamknąć bez zapisu?",
                        parent=self,
                    ):
                        return
                self.destroy()

            self.protocol("WM_DELETE_WINDOW", _confirm_close)

            frm = ttk.Frame(self)
            frm.pack(side="top", fill="both", expand=True, padx=12, pady=(12, 4))
            frm.columnconfigure(1, weight=1)
            frm.columnconfigure(2, weight=1)

            def row_entry(r, label, key):
                ttk.Label(frm, text=label, width=18, anchor="e").grid(
                    row=r, column=0, padx=6, pady=4, sticky="e"
                )
                ent = ttk.Entry(frm, width=36)
                ent.grid(row=r, column=1, padx=6, pady=4, sticky="w")
                ent.insert(0, str(self._row.get(key, "")))
                ent.bind("<KeyRelease>", _mark_dirty)
                return ent

            self.e_id = row_entry(0, "ID / nr_ewid:", "id")
            self.e_nazwa = row_entry(1, "Nazwa:", "nazwa")
            self.e_typ = row_entry(2, "Typ:", "typ")
            self.e_lok = row_entry(3, "Lokalizacja:", "lokalizacja")

            ttk.Label(frm, text="Status:", width=18, anchor="e").grid(
                row=4, column=0, padx=6, pady=4, sticky="e"
            )
            self.cb_status = ttk.Combobox(
                frm, values=self.STATUSES, state="disabled", width=34
            )
            self.cb_status.set(_machine_status_edit_label(self._row.get("status")))
            self.cb_status.grid(row=4, column=1, padx=6, pady=4, sticky="w")
            ttk.Label(
                frm,
                text="Status zmieniaj w trybie Użytkowanie maszyny.",
            ).grid(row=4, column=2, sticky="w", padx=(8, 0))

            def int_or_none(value: str):
                try:
                    return int(value.strip())
                except Exception:
                    return None

            self.e_x = row_entry(5, "x (px):", "x")
            self.e_y = row_entry(6, "y (px):", "y")

            ttk.Label(frm, text="Domyślny typ przeglądu:", width=18, anchor="e").grid(
                row=7, column=0, padx=6, pady=4, sticky="e"
            )
            self.cb_default_review_type = ttk.Combobox(
                frm, values=REVIEW_TYPES, state="readonly", width=34
            )
            self.cb_default_review_type.set(
                str(self._row.get("default_review_type") or REVIEW_TYPES[0])
            )
            self.cb_default_review_type.grid(
                row=7, column=1, padx=6, pady=4, sticky="w"
            )
            self.cb_default_review_type.bind("<<ComboboxSelected>>", _mark_dirty)

            image_frame = ttk.Frame(frm)
            image_frame.grid(
                row=8, column=0, columnspan=3, sticky="w", padx=6, pady=4
            )
            ttk.Button(
                image_frame,
                text="Ustaw zdjęcie...",
                command=self._choose_image,
            ).pack(side="left")
            self.image_label = ttk.Label(image_frame)
            self.image_label.pack(side="left", padx=(12, 0))
            self._refresh_image_label()

            selected_review_months = set(
                _normalize_review_months(
                    self._row.get("review_months")
                    if "review_months" in self._row
                    else self._row.get("review_month")
                )
            )
            self.review_month_vars: Dict[int, tk.BooleanVar] = {}
            review_box = ttk.LabelFrame(frm, text="Przeglądy cykliczne")
            review_box.grid(
                row=9, column=0, columnspan=2, sticky="ew", pady=(10, 4)
            )
            review_box.columnconfigure(1, weight=1)
            ttk.Label(review_box, text="Miesiące przeglądu:").grid(
                row=0, column=0, sticky="nw", padx=6, pady=6
            )
            months_frame = ttk.Frame(review_box)
            months_frame.grid(row=0, column=1, sticky="w", padx=6, pady=6)
            for idx, (month_number, month_label) in enumerate(MONTH_LABELS_PL):
                var = tk.BooleanVar(
                    master=self, value=month_number in selected_review_months
                )
                self.review_month_vars[month_number] = var
                ttk.Checkbutton(
                    months_frame, text=month_label, variable=var
                ).grid(
                    row=idx // 4,
                    column=idx % 4,
                    sticky="w",
                    padx=(0, 12),
                    pady=2,
                )
                var.trace_add("write", _mark_dirty)

            ttk.Label(review_box, text="Dzień miesiąca:").grid(
                row=1, column=0, sticky="e", padx=6, pady=6
            )
            review_day_frame = ttk.Frame(review_box)
            review_day_frame.grid(row=1, column=1, sticky="w", padx=6, pady=6)
            self.review_day_var = tk.StringVar(
                master=self, value=str(_machine_review_day(self._row))
            )
            self.review_day_spin = ttk.Spinbox(
                review_day_frame,
                from_=1,
                to=31,
                width=5,
                textvariable=self.review_day_var,
                state="readonly",
            )
            self.review_day_spin.pack(side="left")
            ttk.Label(
                review_day_frame,
                text="(dla krótszego miesiąca WM użyje ostatniego dnia)",
            ).pack(side="left", padx=(8, 0))
            self.review_day_var.trace_add("write", _mark_dirty)

            ttk.Label(review_box, text="Wykonawcy / serwis:").grid(
                row=2, column=0, sticky="e", padx=6, pady=6
            )
            review_workers = self._row.get("review_workers") or []
            if not isinstance(review_workers, list):
                review_workers = [review_workers]
            selected_workers = {
                str(worker).strip()
                for worker in review_workers
                if str(worker).strip()
            }

            self.review_worker_vars: Dict[str, tk.BooleanVar] = {}
            workers_frame = ttk.Frame(review_box)
            workers_frame.grid(row=2, column=1, sticky="ew", padx=6, pady=6)

            user_logins = _load_wm_user_logins()
            actor = _active_login_for_machine(master)
            if actor and actor not in user_logins:
                user_logins.insert(0, actor)
            for worker in sorted(selected_workers):
                if worker and worker not in user_logins:
                    user_logins.append(worker)

            if not user_logins:
                user_logins = [actor or "system"]

            for idx, login in enumerate(user_logins):
                var = tk.BooleanVar(master=self, value=(login in selected_workers))
                self.review_worker_vars[login] = var
                ttk.Checkbutton(
                    workers_frame,
                    text=login,
                    variable=var,
                ).grid(
                    row=idx // 4,
                    column=idx % 4,
                    sticky="w",
                    padx=(0, 12),
                    pady=2,
                )
                var.trace_add("write", _mark_dirty)

            ttk.Label(
                review_box,
                text=(
                    "Zaznacz sugerowanych serwisantów. Faktycznych wykonawców "
                    "wybierasz przy wykonaniu."
                ),
            ).grid(row=3, column=1, sticky="w", padx=6, pady=(0, 6))

            footer = ttk.Frame(self, padding=(12, 8))
            footer.pack(side="bottom", fill="x")
            ttk.Button(footer, text="Zamknij", command=_confirm_close).pack(
                side="right", padx=(6, 0)
            )
            ttk.Button(footer, text="Anuluj", command=_confirm_close).pack(
                side="right", padx=(6, 0)
            )
            ttk.Button(
                footer,
                text="Zapisz",
                command=lambda: self._ok(
                    int_or_none(self.e_x.get()), int_or_none(self.e_y.get())
                ),
            ).pack(side="right")
            self.bind(
                "<Return>",
                lambda *_: self._ok(int_or_none(self.e_x.get()), int_or_none(self.e_y.get())),
            )
            self.bind("<Escape>", lambda *_: self.destroy())

        def _choose_image(self) -> None:
            path = pick_machine_image(self)
            if not path:
                return
            self._row["image"] = os.path.normpath(path)
            self._refresh_image_label()

        def _refresh_image_label(self) -> None:
            image_path = self._row.get("image") or self._row.get("obraz")
            image_name = os.path.basename(image_path) if image_path else "brak"
            self.image_label.configure(text=f"Zdjęcie: {image_name}")

        def _ok(self, x, y):
            new_status = MACHINE_STATUS_EDIT_LABELS.get(
                self.cb_status.get().strip(), "ok"
            )
            row = {
                "id": (
                    self.e_id.get().strip()
                    or self._row.get("id")
                    or self._row.get("nr_ewid")
                    or ""
                ),
                "nazwa": self.e_nazwa.get().strip(),
                "typ": self.e_typ.get().strip(),
                "lokalizacja": self.e_lok.get().strip(),
                "status": self._old_status,
                "x": x,
                "y": y,
                "default_review_type": (
                    self.cb_default_review_type.get().strip() or REVIEW_TYPES[0]
                ),
                "review_months": [
                    month
                    for month, var in self.review_month_vars.items()
                    if bool(var.get())
                ],
                "review_day": max(1, min(31, int(self.review_day_var.get() or 1))),
                "review_workers": [
                    login
                    for login, var in self.review_worker_vars.items()
                    if bool(var.get())
                ],
            }
            for key in ("status_history", "status_current", "reviews"):
                if key in self._row:
                    row[key] = self._row[key]
            if new_status != self._old_status:
                note = simpledialog.askstring(
                    "Zmiana statusu maszyny",
                    "Opis zmiany statusu:\n"
                    f"{_machine_status_label(self._old_status)} → "
                    f"{_machine_status_label(new_status)}",
                    parent=self,
                )
                if note is None:
                    return
                if new_status in {"alert", "warn"} and not note.strip():
                    messagebox.showwarning(
                        "Maszyny",
                        "Przy zmianie na Serwis / przegląd albo Awarię "
                        "opis jest wymagany.",
                        parent=self,
                    )
                    return
                status_photos: List[str] = []
                try:
                    wants_photos = messagebox.askyesno(
                        "Zdjęcia serwisu / przeglądu",
                        "Czy dodać zdjęcia do tej zmiany statusu?",
                        parent=self,
                    )
                except Exception:
                    wants_photos = False
                if wants_photos:
                    selected_photos = filedialog.askopenfilenames(
                        parent=self,
                        title="Wybierz zdjęcia do historii maszyny",
                        filetypes=[
                            ("Obrazy", "*.png *.jpg *.jpeg *.webp *.bmp"),
                            ("Wszystkie pliki", "*.*"),
                        ],
                    )
                    status_photos = _copy_machine_status_photos(
                        row["id"], selected_photos
                    )
                _apply_machine_status_change(
                    row,
                    new_status,
                    actor=self._actor,
                    note=note.strip(),
                    photos=status_photos,
                )
            else:
                row["status"] = new_status
                _ensure_status_current(row, actor=self._actor)
            row.setdefault("nr_hali", "1")
            if isinstance(self._row.get("zadania"), list):
                row["zadania"] = self._row["zadania"]
            image_path = self._row.get("image") or self._row.get("obraz")
            if image_path:
                norm_path = os.path.normpath(image_path)
                row["image"] = norm_path
                row["obraz"] = norm_path
            if callable(self._on_ok):
                self._on_ok(row)
            self.destroy()

    def _on_add() -> None:
        if not _require_machine_edit_mode():
            return

        def commit(new_row: Dict) -> None:
            nonlocal rows_cache
            if not new_row.get("id"):
                return
            new_row.setdefault("nr_hali", "1")
            new_row.setdefault("zadania", [])
            new_rows = upsert_machine(rows_cache, new_row)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()

        MachineEditDialog(container, row=None, on_ok=commit)

    def _open_machine_usage_window(machine: Dict[str, Any]) -> None:
        """Show the daily machine usage window with status and history."""

        if not machine:
            messagebox.showinfo("Maszyny", "Wybierz maszynę.", parent=root)
            return

        machine_id = str(
            machine.get("id") or machine.get("nr_ewid") or ""
        ).strip()
        if not machine_id:
            messagebox.showwarning(
                "Maszyny",
                "Maszyna nie ma ID / nr_ewid.",
                parent=root,
            )
            return

        win = tk.Toplevel(root)
        win.title(f"Użytkowanie maszyny — {machine_id}")
        win._machine_photo_img = None
        win.geometry("980x720")
        win.minsize(840, 620)
        win.transient(root)

        outer = ttk.Frame(win, padding=12)
        outer.pack(fill="both", expand=True)
        outer.columnconfigure(1, weight=1)
        outer.rowconfigure(2, weight=1)

        photo_frame = ttk.LabelFrame(outer, text="Zdjęcie")
        photo_frame.grid(
            row=0, column=0, rowspan=2, sticky="nsw",
            padx=(0, 12), pady=(0, 8)
        )
        photo_label = ttk.Label(
            photo_frame, text="brak zdjęcia", anchor="center"
        )
        photo_label.configure(width=24)
        photo_label.pack(fill="both", expand=True, padx=8, pady=8)

        def _resolve_machine_image_absolute(path_value: object) -> str:
            raw = str(path_value or "").strip()
            if not raw:
                return ""
            if os.path.isabs(raw) and os.path.exists(raw):
                return os.path.normpath(raw)

            raw_norm = os.path.normpath(raw)
            raw_base = os.path.basename(raw_norm)
            machines_dir = os.path.dirname(primary_path)
            data_root = os.path.dirname(machines_dir)

            candidates: List[str] = []

            # 1) katalog, z którego realnie czytamy maszyny.json
            candidates.extend(
                [
                    os.path.join(machines_dir, raw_norm),
                    os.path.join(machines_dir, raw_base),
                    os.path.join(machines_dir, "zdjecia", raw_norm),
                    os.path.join(machines_dir, "zdjecia", raw_base),
                    os.path.join(machines_dir, "zdjęcia", raw_norm),
                    os.path.join(machines_dir, "zdjęcia", raw_base),
                    os.path.join(machines_dir, "images", raw_norm),
                    os.path.join(machines_dir, "images", raw_base),
                    os.path.join(machines_dir, "photos", raw_norm),
                    os.path.join(machines_dir, "photos", raw_base),
                    os.path.join(
                        machines_dir, "attachments", machine_id, raw_norm
                    ),
                    os.path.join(
                        machines_dir, "attachments", machine_id, raw_base
                    ),
                ]
            )

            # 2) ROOT/data jako fallback
            candidates.extend(
                [
                    os.path.join(data_root, "maszyny", raw_norm),
                    os.path.join(data_root, "maszyny", raw_base),
                    os.path.join(data_root, "maszyny", "zdjecia", raw_norm),
                    os.path.join(data_root, "maszyny", "zdjecia", raw_base),
                    os.path.join(data_root, "maszyny", "zdjęcia", raw_norm),
                    os.path.join(data_root, "maszyny", "zdjęcia", raw_base),
                    os.path.join(data_root, "maszyny", "images", raw_norm),
                    os.path.join(data_root, "maszyny", "images", raw_base),
                    os.path.join(data_root, "maszyny", "photos", raw_norm),
                    os.path.join(data_root, "maszyny", "photos", raw_base),
                    os.path.join(
                        data_root, "maszyny", "attachments", machine_id,
                        raw_norm
                    ),
                    os.path.join(
                        data_root, "maszyny", "attachments", machine_id,
                        raw_base
                    ),
                ]
            )

            try:
                cfg_data_root = (
                    cfg_manager.path_data() if cfg_manager is not None else ""
                )
                if cfg_data_root:
                    candidates.extend(
                        [
                            os.path.join(cfg_data_root, "maszyny", raw_norm),
                            os.path.join(cfg_data_root, "maszyny", raw_base),
                            os.path.join(
                                cfg_data_root, "maszyny", "zdjecia", raw_norm
                            ),
                            os.path.join(
                                cfg_data_root, "maszyny", "zdjecia", raw_base
                            ),
                            os.path.join(
                                cfg_data_root, "maszyny", "zdjęcia", raw_norm
                            ),
                            os.path.join(
                                cfg_data_root, "maszyny", "zdjęcia", raw_base
                            ),
                            os.path.join(
                                cfg_data_root, "maszyny", "images", raw_norm
                            ),
                            os.path.join(
                                cfg_data_root, "maszyny", "images", raw_base
                            ),
                            os.path.join(
                                cfg_data_root, "maszyny", "photos", raw_norm
                            ),
                            os.path.join(
                                cfg_data_root, "maszyny", "photos", raw_base
                            ),
                            os.path.join(
                                cfg_data_root,
                                "maszyny",
                                "attachments",
                                machine_id,
                                raw_norm,
                            ),
                            os.path.join(
                                cfg_data_root,
                                "maszyny",
                                "attachments",
                                machine_id,
                                raw_base,
                            ),
                        ]
                    )
            except Exception:
                pass

            candidates.append(raw_norm)

            seen: set[str] = set()
            for candidate in candidates:
                candidate = os.path.normpath(candidate)
                if candidate in seen:
                    continue
                seen.add(candidate)
                if os.path.exists(candidate):
                    return candidate

            # 3) ostatnia deska ratunku: szybkie szukanie po nazwie pliku
            #    w data/maszyny. Przydatne, gdy w JSON jest samo
            #    "100004526.jpg", a folder zdjęć jest inny.
            if raw_base:
                search_roots = [
                    machines_dir,
                    os.path.join(data_root, "maszyny"),
                ]
                for search_root in search_roots:
                    if not search_root or not os.path.isdir(search_root):
                        continue
                    try:
                        for dirpath, _dirnames, filenames in os.walk(
                            search_root
                        ):
                            if raw_base in filenames:
                                return os.path.normpath(
                                    os.path.join(dirpath, raw_base)
                                )
                    except Exception:
                        pass
            return ""

        image_path = (
            machine.get("image") or machine.get("obraz")
            or machine.get("photo")
        )
        if image_path:
            try:
                resolved_image = _resolve_machine_image_absolute(image_path)
                image_exists = bool(
                    resolved_image and os.path.exists(resolved_image)
                )
                try:
                    print(
                        "[WM-DBG][MASZYNY][PHOTO] "
                        f"id={machine_id} raw={image_path!r} "
                        f"resolved={resolved_image!r} "
                        f"exists={int(image_exists)} "
                        f"pil={int(bool(Image and ImageTk))}"
                    )
                except Exception:
                    pass
                if image_exists and Image and ImageTk:
                    img = Image.open(resolved_image)
                    img.thumbnail((180, 180), Image.LANCZOS)
                    photo_img = ImageTk.PhotoImage(img)
                    win._machine_photo_img = photo_img
                    photo_label.configure(image=photo_img, text="")
                elif image_exists:
                    photo_label.configure(
                        text=os.path.basename(resolved_image)
                    )
                else:
                    photo_label.configure(text=f"Nie znaleziono:\n{image_path}")
            except Exception as exc:
                try:
                    print(
                        "[WM-DBG][MASZYNY][PHOTO][ERR] "
                        f"id={machine_id} raw={image_path!r} error={exc}"
                    )
                except Exception:
                    pass
                photo_label.configure(text=str(image_path))

        header = ttk.Frame(outer)
        header.grid(row=0, column=1, sticky="new")
        header.columnconfigure(1, weight=1)

        ttk.Label(
            header, text=machine_id, font=("TkDefaultFont", 28, "bold")
        ).grid(row=0, column=0, sticky="w", padx=(0, 16))
        title_text = f"{machine.get('nazwa') or machine.get('name') or ''}"
        typ_text = f"{machine.get('typ') or machine.get('type') or ''}"
        ttk.Label(
            header, text=title_text, font=("TkDefaultFont", 18, "bold")
        ).grid(row=0, column=1, sticky="w")
        ttk.Label(header, text=typ_text).grid(
            row=1, column=1, sticky="w", pady=(2, 0)
        )

        status_key = _normalize_machine_status(machine.get("status"))
        status_label = _machine_status_label(status_key)
        status_box = ttk.LabelFrame(outer, text="Aktualny status")
        status_box.grid(row=1, column=1, sticky="ew", pady=(8, 8))
        ttk.Label(
            status_box,
            text=status_label,
            foreground=_machine_status_color(status_key),
            font=("TkDefaultFont", 24, "bold"),
        ).pack(anchor="w", padx=10, pady=8)

        summary = machine.get("__schedule_summary") or {}
        next_review = str(summary.get("next_label") or "—")
        review_status = str(
            summary.get("status_text") or "Brak danych przeglądu"
        )
        review_box = ttk.LabelFrame(outer, text="Najbliższy przegląd")
        review_box.grid(
            row=2, column=0, columnspan=2, sticky="new", pady=(0, 8)
        )
        ttk.Label(review_box, text=f"Termin: {next_review}").pack(
            anchor="w", padx=8, pady=(6, 2)
        )
        ttk.Label(review_box, text=review_status).pack(
            anchor="w", padx=8, pady=(0, 6)
        )

        history_box = ttk.LabelFrame(outer, text="Historia statusów")
        history_box.grid(
            row=3, column=0, columnspan=2, sticky="nsew", pady=(0, 8)
        )
        outer.rowconfigure(3, weight=1)
        hist_cols = ("status", "start", "stop", "kto", "zdarzenie")
        hist_tree = ttk.Treeview(
            history_box, columns=hist_cols, show="headings", height=9
        )
        hist_setup = {
            "status": ("Status", 135, "w"),
            "start": ("Start", 155, "center"),
            "stop": ("Stop", 155, "center"),
            "kto": ("Kto", 120, "w"),
            "zdarzenie": ("Powód / zdarzenie", 430, "w"),
        }
        for col, (label, width, anchor) in hist_setup.items():
            hist_tree.heading(col, text=label)
            hist_tree.column(col, width=width, anchor=anchor)
        hist_tree.pack(fill="both", expand=True, padx=6, pady=6)

        history_items: Dict[str, Dict[str, Any]] = {}

        def _history_entries_for_usage(
            src: Dict[str, Any]
        ) -> List[Dict[str, Any]]:
            entries: List[Dict[str, Any]] = []
            raw_history = src.get("status_history")
            if isinstance(raw_history, list):
                for item in raw_history:
                    if isinstance(item, dict):
                        entries.append(dict(item))
            current = src.get("status_current")
            if isinstance(current, dict):
                item = dict(current)
                item["__current"] = True
                entries.append(item)
            return entries

        def _history_entry_values(item: Dict[str, Any]) -> tuple:
            status = _machine_status_label(item.get("status"))
            start = _format_machine_history_dt(item.get("started_at"))
            if item.get("__current"):
                stop = "w toku"
                who = str(item.get("changed_by") or "—")
                note = str(item.get("note") or "")
            else:
                stop = _format_machine_history_dt(item.get("ended_at"))
                who = str(item.get("closed_by") or item.get("changed_by") or "—")
                note = str(item.get("close_note") or item.get("note") or "")
            photos = (
                item.get("photos") if isinstance(item.get("photos"), list) else []
            )
            event_text = note.strip() or "Zmiana statusu"
            if photos:
                event_text += f" | zdjęcia: {len(photos)}"
            return status, start, stop, who, event_text

        def _refresh_history_tree() -> None:
            history_items.clear()
            for iid in hist_tree.get_children():
                hist_tree.delete(iid)
            entries = _history_entries_for_usage(machine)
            if entries:
                for item in entries:
                    iid = hist_tree.insert(
                        "", "end", values=_history_entry_values(item)
                    )
                    history_items[iid] = item
            else:
                hist_tree.insert(
                    "",
                    "end",
                    values=(
                        "—",
                        "—",
                        "—",
                        "—",
                        "Brak historii. Pierwszy wpis powstanie przy zmianie statusu.",
                    ),
                )

        _refresh_history_tree()

        def _selected_history_item() -> Optional[Dict[str, Any]]:
            selected = hist_tree.selection()
            if not selected:
                return None
            return history_items.get(selected[0])

        def _show_history_photos() -> None:
            item = _selected_history_item()
            if not item:
                messagebox.showinfo(
                    "Zdjęcia historii",
                    "Wybierz wpis historii.",
                    parent=win,
                )
                return
            photos = (
                item.get("photos") if isinstance(item.get("photos"), list) else []
            )
            if not photos:
                messagebox.showinfo(
                    "Zdjęcia historii",
                    "Ten wpis historii nie ma zdjęć.",
                    parent=win,
                )
                return

            photo_win = tk.Toplevel(win)
            photo_win.title("Zdjęcia z historii maszyny")
            photo_win.geometry("760x520")
            photo_win.minsize(640, 420)
            photo_win.transient(win)
            photo_win._thumbs = []

            header_text = (
                f"{_machine_status_label(item.get('status'))} | "
                f"{str(item.get('started_at') or '—').replace('T', ' ')[:16]}"
            )
            ttk.Label(
                photo_win,
                text=header_text,
                font=("TkDefaultFont", 14, "bold"),
            ).pack(anchor="w", padx=10, pady=(10, 4))

            holder = ttk.Frame(photo_win, padding=8)
            holder.pack(fill="both", expand=True)

            for idx, raw_path in enumerate(photos, start=1):
                resolved = _resolve_machine_image_absolute(raw_path)
                if not resolved:
                    resolved = _resolve_card_absolute(str(raw_path), cfg_manager)
                exists = bool(resolved and os.path.exists(resolved))

                tile = ttk.LabelFrame(holder, text=f"Zdjęcie {idx}")
                tile.grid(
                    row=(idx - 1) // 4,
                    column=(idx - 1) % 4,
                    padx=6,
                    pady=6,
                    sticky="nsew",
                )

                if exists and Image and ImageTk:
                    try:
                        img = Image.open(resolved)
                        img.thumbnail((140, 140), Image.LANCZOS)
                        thumb = ImageTk.PhotoImage(img)
                        photo_win._thumbs.append(thumb)
                        ttk.Button(
                            tile,
                            image=thumb,
                            command=lambda p=resolved: _open_external(p),
                        ).pack(padx=6, pady=6)
                    except Exception:
                        ttk.Button(
                            tile,
                            text=os.path.basename(resolved),
                            command=lambda p=resolved: _open_external(p),
                        ).pack(padx=6, pady=6)
                elif exists:
                    ttk.Button(
                        tile,
                        text=os.path.basename(resolved),
                        command=lambda p=resolved: _open_external(p),
                    ).pack(padx=6, pady=6)
                else:
                    ttk.Label(
                        tile,
                        text=f"Nie znaleziono:\n{raw_path}",
                        wraplength=150,
                    ).pack(padx=6, pady=6)

            ttk.Button(
                photo_win, text="Zamknij", command=photo_win.destroy
            ).pack(anchor="e", padx=10, pady=(0, 10))

        hist_tree.bind("<Double-1>", lambda _event: _show_history_photos())

        history_actions = ttk.Frame(history_box)
        history_actions.pack(fill="x", padx=6, pady=(0, 6))
        ttk.Button(
            history_actions,
            text="Pokaż zdjęcia",
            command=_show_history_photos,
        ).pack(side="left")

        reviews_box = ttk.LabelFrame(outer, text="Przeglądy / serwis maszyny")
        reviews_box.grid(
            row=4, column=0, columnspan=2, sticky="nsew", pady=(0, 8)
        )

        reviews_cols = ("date", "type", "status", "people", "details")
        reviews_tree = ttk.Treeview(
            reviews_box,
            columns=reviews_cols,
            show="headings",
            height=6,
        )
        reviews_setup = {
            "date": ("Data", 105, "center"),
            "type": ("Typ", 165, "w"),
            "status": ("Status", 105, "center"),
            "people": ("Osoby", 190, "w"),
            "details": ("Szczegóły", 420, "w"),
        }
        for col, (label, width, anchor) in reviews_setup.items():
            reviews_tree.heading(col, text=label)
            reviews_tree.column(col, width=width, anchor=anchor)
        reviews_tree.pack(fill="both", expand=True, padx=6, pady=6)

        review_items: Dict[str, Dict[str, Any]] = {}

        def _review_status_label(value: object) -> str:
            raw = str(value or "").strip().lower()
            if raw in {"done", "wykonany", "completed"}:
                return "Wykonany"
            if raw in {"in_progress", "w_trakcie", "w trakcie"}:
                return "W trakcie"
            return "Planowany"

        def _people_text(value: object) -> str:
            if isinstance(value, list):
                return ", ".join(
                    str(item) for item in value if str(item).strip()
                )
            return str(value or "")

        def _linked_dysp_id(entry: Dict[str, Any]) -> str:
            direct = str(entry.get("dyspozycja_id") or "").strip()
            if direct:
                return direct
            source = str(entry.get("source") or "").strip().lower()
            if source != REVIEW_SOURCE_CYCLE:
                return ""
            try:
                linked = find_cycle_dyspozycja_for_review(machine, entry)
            except Exception:
                linked = None
            return str((linked or {}).get("id") or "").strip()

        def _refresh_reviews_tree() -> None:
            review_items.clear()
            for iid in reviews_tree.get_children():
                reviews_tree.delete(iid)
            entries = _combined_machine_review_entries(
                machine, today=dt.date.today(), years_ahead=1
            )
            month_names = dict(MONTH_LABELS_PL)
            for entry in entries:
                source = str(entry.get("source") or REVIEW_SOURCE_MANUAL).strip().lower()
                is_cycle = source == REVIEW_SOURCE_CYCLE
                date_value = _review_date(
                    entry.get("date")
                    or entry.get("planned_date")
                    or entry.get("completed_at")
                )
                planned_text = (
                    _format_machine_review_date(date_value)
                    if date_value is not None
                    else str(entry.get("planned_date") or "—")
                )
                type_text = (
                    "Przegląd cykliczny"
                    if is_cycle
                    else str(entry.get("type") or "")
                )
                cycle_text = ""
                if is_cycle and date_value is not None:
                    cycle_text = (
                        f"Cykliczny: {month_names.get(date_value.month, str(date_value.month))} "
                        f"{date_value.year}"
                    )

                status_label = _review_status_label(entry.get("status"))
                if status_label == "Wykonany":
                    people = _people_text(entry.get("completed_by"))
                    completed_at = _format_machine_history_dt(entry.get("completed_at")) if entry.get("completed_at") else ""
                    details = str(entry.get("result_note") or entry.get("description") or "")
                    if completed_at:
                        details = f"Wykonano: {completed_at}" + (f" | {details}" if details else "")
                elif status_label == "W trakcie":
                    people = str(entry.get("started_by") or "") or _people_text(
                        entry.get("suggested_workers") or entry.get("suggested_people")
                    )
                    started_at = _format_machine_history_dt(entry.get("started_at")) if entry.get("started_at") else ""
                    details = str(entry.get("description") or "")
                    if started_at:
                        details = f"Rozpoczęto: {started_at}" + (f" | {details}" if details else "")
                else:
                    people = _people_text(
                        entry.get("suggested_workers") or entry.get("suggested_people")
                    )
                    details = str(entry.get("description") or "")

                if cycle_text and cycle_text.lower() not in details.lower():
                    details = cycle_text + (f" | {details}" if details else "")
                dysp_id = _linked_dysp_id(entry)
                if dysp_id and dysp_id.lower() not in details.lower():
                    details = (details + " | " if details else "") + f"Dyspozycja: {dysp_id}"

                values = (
                    planned_text,
                    type_text,
                    status_label,
                    people or "—",
                    details or "—",
                )
                iid = reviews_tree.insert("", "end", values=values)
                review_items[iid] = entry

        def _selected_review_entry() -> Optional[Dict[str, Any]]:
            sel = reviews_tree.selection()
            if not sel:
                return None
            return review_items.get(sel[0])

        def _show_selected_review_details(_event=None) -> None:
            entry = _selected_review_entry()
            if not entry:
                return
            source = str(entry.get("source") or REVIEW_SOURCE_MANUAL).strip().lower()
            source_label = REVIEW_SOURCE_LABELS.get(source, source or "Ręczny")
            planned = _review_date(
                entry.get("date") or entry.get("planned_date") or entry.get("completed_at")
            )
            lines = [
                f"Maszyna: {machine_id} — {machine.get('nazwa') or machine.get('name') or ''}",
                f"Plan: {_format_machine_review_date(planned) if planned else '—'}",
                f"Typ: {entry.get('type') or 'Przegląd okresowy'}",
                f"Źródło: {source_label}",
                f"Status: {_review_status_label(entry.get('status'))}",
                f"Dyspozycja: {_linked_dysp_id(entry) or '—'}",
                f"Rozpoczął: {entry.get('started_by') or '—'}",
                f"Start: {_format_machine_history_dt(entry.get('started_at')) if entry.get('started_at') else '—'}",
                f"Wykonali: {_people_text(entry.get('completed_by')) or '—'}",
                f"Wykonano: {_format_machine_history_dt(entry.get('completed_at')) if entry.get('completed_at') else '—'}",
                f"Zakres / opis: {entry.get('description') or '—'}",
                f"Wynik / uwagi: {entry.get('result_note') or '—'}",
                f"Zdjęcia: {len(entry.get('photos') or []) if isinstance(entry.get('photos'), list) else 0}",
            ]
            messagebox.showinfo(
                "Karta serwisowa maszyny",
                "\n".join(lines),
                parent=win,
            )

        def _review_for_action(display_entry: Dict[str, Any]) -> Dict[str, Any]:
            wanted_id = str(display_entry.get("id") or "").strip()
            source = str(display_entry.get("source") or REVIEW_SOURCE_MANUAL).strip().lower()
            wanted_date = _review_date(
                display_entry.get("date")
                or display_entry.get("planned_date")
                or display_entry.get("completed_at")
            )
            wanted_type = str(display_entry.get("type") or "Przegląd okresowy").strip()

            for current in _machine_reviews(machine):
                if wanted_id and str(current.get("id") or "").strip() == wanted_id:
                    return current

            if source != REVIEW_SOURCE_CYCLE or not wanted_id.startswith("cycle_"):
                for current in _machine_reviews(machine):
                    current_date = _review_date(
                        current.get("date")
                        or current.get("planned_date")
                        or current.get("completed_at")
                    )
                    current_type = str(current.get("type") or "Przegląd okresowy").strip()
                    if current_date == wanted_date and current_type == wanted_type:
                        return current
                return display_entry

            if wanted_date is None:
                return display_entry

            for current in _machine_reviews(machine):
                current_date = _review_date(
                    current.get("date")
                    or current.get("planned_date")
                    or current.get("completed_at")
                )
                current_type = str(current.get("type") or "Przegląd okresowy").strip()
                current_source = str(current.get("source") or "").strip().lower()
                if (
                    current_source == REVIEW_SOURCE_CYCLE
                    and current_date is not None
                    and current_date.year == wanted_date.year
                    and current_date.month == wanted_date.month
                    and current_type == wanted_type
                ):
                    return current

            month_name = dict(MONTH_LABELS_PL).get(wanted_date.month, str(wanted_date.month))
            persisted = {
                "id": _new_review_id(),
                "type": wanted_type or "Przegląd okresowy",
                "planned_date": wanted_date.isoformat(),
                "status": REVIEW_STATUS_PLANNED,
                "source": REVIEW_SOURCE_CYCLE,
                "cycle_year": wanted_date.year,
                "cycle_month": wanted_date.month,
                "suggested_workers": list(
                    display_entry.get("suggested_workers")
                    or display_entry.get("suggested_people")
                    or []
                ),
                "description": f"Przegląd cykliczny: {month_name} {wanted_date.year}",
                "completed_at": "",
                "completed_by": [],
                "result_note": "",
                "photos": [],
            }
            reviews = list(_machine_reviews(machine))
            reviews.append(persisted)
            machine["reviews"] = reviews
            return persisted

        def _persist_machine_after_review_change(
            updated_machine: Dict[str, Any],
        ) -> None:
            nonlocal rows_cache
            new_rows = upsert_machine(rows_cache, updated_machine)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()
            _set_selected_machine(machine_id)

        def _start_selected_review() -> None:
            entry = _selected_review_entry()
            if not entry:
                messagebox.showinfo("Przegląd / serwis", "Wybierz wpis.", parent=win)
                return
            if str(entry.get("status") or "").lower() in {
                "done", "wykonany", "completed"
            }:
                messagebox.showinfo(
                    "Przegląd / serwis",
                    "Ten wpis jest już wykonany.",
                    parent=win,
                )
                return
            if _normalize_machine_status(machine.get("status")) == "warn":
                messagebox.showwarning(
                    "Przegląd / serwis",
                    "Maszyna jest w statusie Awaria. Najpierw zamknij "
                    "awarię albo zmień status ręcznie.",
                    parent=win,
                )
                return

            entry = _review_for_action(entry)

            note = (
                f"Rozpoczęto {entry.get('type') or 'przegląd / serwis'}"
                f" | plan: {entry.get('planned_date') or '—'}"
            )
            if entry.get("description"):
                note += f" | {entry.get('description')}"

            actor = _active_login_for_machine(root)
            entry["status"] = "in_progress"
            entry["started_at"] = _machine_now_iso()
            entry["started_by"] = actor
            try:
                linked = sync_review_to_dyspozycja(
                    machine,
                    entry,
                    status="in_progress",
                    actor=actor,
                    note=note,
                )
            except Exception:
                logger.exception(
                    "[Maszyny][DYSP] Nie udało się rozpocząć powiązanej Dyspozycji."
                )
                linked = None
            if linked:
                entry["dyspozycja_id"] = str(linked.get("id") or "")
                note += f" | Dyspozycja: {entry['dyspozycja_id']}"

            updated = dict(machine)
            updated["reviews"] = list(_machine_reviews(machine))
            _apply_machine_status_change(
                updated,
                "alert",
                actor=actor,
                note=note,
                photos=[],
            )
            machine.update(updated)
            _persist_machine_after_review_change(updated)
            _refresh_history_tree()
            _refresh_reviews_tree()
            try:
                root.event_generate("<<DyspozycjeUpdated>>", when="tail")
            except Exception:
                pass

        def _persist_machine_reviews() -> None:
            updated = dict(machine)
            updated["reviews"] = list(_machine_reviews(machine))
            _persist_machine_after_review_change(updated)

        def _open_add_review_dialog() -> None:
            dialog = tk.Toplevel(win)
            dialog.title("Dodaj przegląd / serwis")
            dialog.geometry("620x420")
            dialog.transient(win)
            dialog.grab_set()

            frm = ttk.Frame(dialog, padding=12)
            frm.pack(fill="both", expand=True)
            frm.columnconfigure(1, weight=1)

            ttk.Label(frm, text="Typ:").grid(
                row=0, column=0, sticky="e", padx=4, pady=4
            )
            default_review_type = str(
                machine.get("default_review_type") or REVIEW_TYPES[0]
            )
            var_type = tk.StringVar(value=default_review_type)
            ttk.Combobox(
                frm,
                textvariable=var_type,
                values=REVIEW_TYPES,
                state="readonly",
            ).grid(row=0, column=1, sticky="ew", padx=4, pady=4)

            ttk.Label(frm, text="Planowana data:").grid(
                row=1, column=0, sticky="e", padx=4, pady=4
            )
            var_date = tk.StringVar(value=dt.date.today().isoformat())
            ttk.Entry(frm, textvariable=var_date).grid(
                row=1, column=1, sticky="ew", padx=4, pady=4
            )

            ttk.Label(frm, text="Sugerowani:").grid(
                row=2, column=0, sticky="e", padx=4, pady=4
            )
            suggested = machine.get("review_workers")
            if isinstance(suggested, list):
                suggested_text = ", ".join(
                    str(x) for x in suggested if str(x).strip()
                )
            else:
                suggested_text = str(suggested or "")
            var_suggested = tk.StringVar(value=suggested_text)
            ttk.Entry(frm, textvariable=var_suggested).grid(
                row=2, column=1, sticky="ew", padx=4, pady=4
            )
            ttk.Label(
                frm,
                text=(
                    "To tylko sugestia. Faktycznych wykonawców wybierasz "
                    "dopiero przy wykonaniu."
                ),
            ).grid(row=3, column=1, sticky="w", padx=4, pady=(0, 6))

            ttk.Label(frm, text="Zakres / opis:").grid(
                row=4, column=0, sticky="ne", padx=4, pady=4
            )
            txt_desc = tk.Text(frm, height=8, wrap="word")
            txt_desc.grid(row=4, column=1, sticky="nsew", padx=4, pady=4)
            frm.rowconfigure(4, weight=1)

            def _save_review() -> None:
                parsed = _parse_schedule_date(var_date.get().strip())
                if parsed is None:
                    messagebox.showwarning(
                        "Przegląd / serwis",
                        "Podaj poprawną datę, np. 2026-06-01.",
                        parent=dialog,
                    )
                    return
                entry = {
                    "id": _new_review_id(),
                    "type": var_type.get().strip() or REVIEW_TYPES[0],
                    "planned_date": parsed.isoformat(),
                    "status": "planned",
                    "suggested_workers": _split_csv_people(var_suggested.get()),
                    "description": txt_desc.get("1.0", "end").strip(),
                    "completed_at": "",
                    "completed_by": [],
                    "result_note": "",
                    "photos": [],
                }
                reviews = list(_machine_reviews(machine))
                reviews.append(entry)
                machine["reviews"] = reviews
                _persist_machine_reviews()
                _refresh_reviews_tree()
                dialog.destroy()

            btns = ttk.Frame(frm)
            btns.grid(row=5, column=0, columnspan=2, sticky="e", pady=(10, 0))
            ttk.Button(btns, text="Zapisz", command=_save_review).pack(
                side="left", padx=4
            )
            ttk.Button(btns, text="Anuluj", command=dialog.destroy).pack(
                side="left", padx=4
            )

        def _open_complete_review_dialog() -> None:
            entry = _selected_review_entry()
            if not entry:
                messagebox.showinfo("Przegląd / serwis", "Wybierz wpis.", parent=win)
                return
            if str(entry.get("status") or "").lower() in {
                "done", "wykonany", "completed"
            }:
                messagebox.showinfo(
                    "Przegląd / serwis", "Ten wpis jest już wykonany.", parent=win
                )
                return

            dialog = tk.Toplevel(win)
            dialog.title("Oznacz przegląd / serwis jako wykonany")
            dialog.geometry("620x520")
            dialog.transient(win)
            dialog.grab_set()

            frm = ttk.Frame(dialog, padding=12)
            frm.pack(fill="both", expand=True)
            frm.columnconfigure(1, weight=1)

            ttk.Label(frm, text="Wykonali:").grid(
                row=0, column=0, sticky="ne", padx=4, pady=4
            )
            users_box = ttk.Frame(frm)
            users_box.grid(row=0, column=1, sticky="ew", padx=4, pady=4)

            user_logins = _load_wm_user_logins()
            actor = _active_login_for_machine(root)
            if actor and actor not in user_logins:
                user_logins.insert(0, actor)
            if not user_logins:
                user_logins = [actor or "system"]

            selected_vars: Dict[str, tk.BooleanVar] = {}
            for idx, login in enumerate(user_logins):
                var = tk.BooleanVar(value=(login == actor))
                selected_vars[login] = var
                ttk.Checkbutton(
                    users_box,
                    text=login,
                    variable=var,
                ).grid(
                    row=idx // 3,
                    column=idx % 3,
                    sticky="w",
                    padx=(0, 12),
                    pady=2,
                )

            ttk.Label(frm, text="Co wykonano:").grid(
                row=1, column=0, sticky="ne", padx=4, pady=4
            )
            txt_result = tk.Text(frm, height=10, wrap="word")
            txt_result.grid(row=1, column=1, sticky="nsew", padx=4, pady=4)
            frm.rowconfigure(1, weight=1)

            def _save_completed() -> None:
                completed_by = [
                    login for login, var in selected_vars.items() if bool(var.get())
                ]
                if not completed_by:
                    messagebox.showwarning(
                        "Przegląd / serwis",
                        "Wybierz przynajmniej jedną osobę, która wykonała "
                        "przegląd/serwis.",
                        parent=dialog,
                    )
                    return
                target_entry = _review_for_action(entry)
                target_entry["status"] = "done"
                target_entry["completed_at"] = _machine_now_iso()
                target_entry["completed_by"] = completed_by
                target_entry["result_note"] = txt_result.get("1.0", "end").strip()

                actor = ", ".join(completed_by)
                note = (
                    f"Wykonano {target_entry.get('type') or 'przegląd / serwis'}"
                    f" | plan: {target_entry.get('planned_date') or '—'}"
                )
                if target_entry.get("result_note"):
                    note += f" | {target_entry.get('result_note')}"
                try:
                    linked = sync_review_to_dyspozycja(
                        machine,
                        target_entry,
                        status="done",
                        actor=actor,
                        note=target_entry.get("result_note") or note,
                    )
                except Exception:
                    logger.exception(
                        "[Maszyny][DYSP] Nie udało się zamknąć powiązanej Dyspozycji."
                    )
                    linked = None
                if linked:
                    target_entry["dyspozycja_id"] = str(linked.get("id") or "")
                    note += f" | Dyspozycja: {target_entry['dyspozycja_id']}"

                updated = dict(machine)
                updated["reviews"] = list(_machine_reviews(machine))
                if _normalize_machine_status(updated.get("status")) == "alert":
                    _apply_machine_status_change(
                        updated,
                        "ok",
                        actor=actor,
                        note=note,
                        photos=[],
                    )
                machine.update(updated)
                _persist_machine_after_review_change(updated)
                _refresh_history_tree()
                _refresh_reviews_tree()
                try:
                    root.event_generate("<<DyspozycjeUpdated>>", when="tail")
                except Exception:
                    pass
                dialog.destroy()

            btns = ttk.Frame(frm)
            btns.grid(row=2, column=0, columnspan=2, sticky="e", pady=(10, 0))
            ttk.Button(
                btns, text="Zapisz wykonanie", command=_save_completed
            ).pack(side="left", padx=4)
            ttk.Button(btns, text="Anuluj", command=dialog.destroy).pack(
                side="left", padx=4
            )

        _refresh_reviews_tree()
        reviews_tree.bind("<Double-1>", _show_selected_review_details, add=True)

        reviews_actions = ttk.Frame(reviews_box)
        reviews_actions.pack(fill="x", padx=6, pady=(0, 6))
        ttk.Button(
            reviews_actions,
            text="Dodaj przegląd / serwis",
            command=_open_add_review_dialog,
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            reviews_actions,
            text="Rozpocznij przegląd / serwis",
            command=_start_selected_review,
        ).pack(side="left", padx=(0, 6))
        ttk.Button(
            reviews_actions,
            text="Oznacz jako wykonany",
            command=_open_complete_review_dialog,
        ).pack(side="left")

        buttons = ttk.Frame(outer)
        buttons.grid(row=5, column=0, columnspan=2, sticky="e")
        ttk.Button(
            buttons,
            text="Zmień status",
            command=lambda: (
                win.destroy(), _open_status_change_dialog(machine)
            ),
        ).pack(side="left", padx=4)
        ttk.Button(
            buttons, text="Zamknij", command=win.destroy
        ).pack(side="left", padx=4)

    def _open_status_change_dialog(machine: Dict[str, Any]) -> None:
        """Change only the operational status, without technical editing."""

        nonlocal rows_cache

        machine_id = str(
            machine.get("id") or machine.get("nr_ewid") or ""
        ).strip()
        if not machine_id:
            messagebox.showwarning(
                "Maszyny",
                "Maszyna nie ma ID / nr_ewid.",
                parent=root,
            )
            return

        old_status = _normalize_machine_status(machine.get("status"))
        win = tk.Toplevel(root)
        win.title("Zmień status maszyny")
        win.geometry("520x260")
        win.transient(root)
        win.grab_set()

        box = ttk.Frame(win, padding=12)
        box.pack(fill="both", expand=True)
        box.columnconfigure(1, weight=1)
        box.rowconfigure(3, weight=1)

        ttk.Label(box, text="Maszyna:").grid(
            row=0, column=0, sticky="e", padx=4, pady=4
        )
        ttk.Label(
            box,
            text=(
                f"{machine_id} — "
                f"{machine.get('nazwa') or machine.get('name') or ''}"
            ),
        ).grid(row=0, column=1, sticky="w", padx=4, pady=4)
        ttk.Label(box, text="Aktualny status:").grid(
            row=1, column=0, sticky="e", padx=4, pady=4
        )
        ttk.Label(box, text=_machine_status_label(old_status)).grid(
            row=1, column=1, sticky="w", padx=4, pady=4
        )
        ttk.Label(box, text="Nowy status:").grid(
            row=2, column=0, sticky="e", padx=4, pady=4
        )
        status_var = tk.StringVar(
            value=_machine_status_edit_label(old_status)
        )
        status_box = ttk.Combobox(
            box,
            textvariable=status_var,
            values=MACHINE_STATUS_EDIT_VALUES,
            state="readonly",
            width=32,
        )
        status_box.grid(row=2, column=1, sticky="ew", padx=4, pady=4)
        ttk.Label(box, text="Opis:").grid(
            row=3, column=0, sticky="ne", padx=4, pady=4
        )
        note_text = tk.Text(box, height=5, wrap="word")
        note_text.grid(row=3, column=1, sticky="nsew", padx=4, pady=4)

        def _save_status() -> None:
            nonlocal rows_cache

            new_status = MACHINE_STATUS_EDIT_LABELS.get(
                status_var.get().strip(), "ok"
            )
            note = note_text.get("1.0", "end").strip()
            if new_status == old_status:
                win.destroy()
                return
            if new_status in {"alert", "warn"} and not note:
                messagebox.showwarning(
                    "Maszyny",
                    "Przy zmianie na Serwis / przegląd albo Awarię "
                    "opis jest wymagany.",
                    parent=win,
                )
                return

            status_photos: List[str] = []
            try:
                wants_photos = messagebox.askyesno(
                    "Zdjęcia serwisu / przeglądu",
                    "Czy dodać zdjęcia do tej zmiany statusu?",
                    parent=win,
                )
            except Exception:
                wants_photos = False
            if wants_photos:
                selected_photos = filedialog.askopenfilenames(
                    parent=win,
                    title="Wybierz zdjęcia do historii maszyny",
                    filetypes=[
                        ("Obrazy", "*.png *.jpg *.jpeg *.webp *.bmp"),
                        ("Wszystkie pliki", "*.*"),
                    ],
                )
                status_photos = _copy_machine_status_photos(
                    machine_id, selected_photos
                )

            payload = dict(machine)
            payload["status"] = old_status
            _apply_machine_status_change(
                payload,
                new_status,
                actor=_active_login_for_machine(root),
                note=note,
                photos=status_photos,
            )
            new_rows = upsert_machine(rows_cache, payload)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()
            _set_selected_machine(machine_id)
            win.destroy()

        buttons = ttk.Frame(box)
        buttons.grid(
            row=4, column=0, columnspan=2, sticky="e", pady=(10, 0)
        )
        ttk.Button(buttons, text="Zapisz", command=_save_status).pack(
            side="left", padx=4
        )
        ttk.Button(buttons, text="Anuluj", command=win.destroy).pack(
            side="left", padx=4
        )

    def _on_change_status() -> None:
        machine_id = _selected_id()
        machine = _find_machine(machine_id) if machine_id else None
        if not machine:
            messagebox.showinfo(
                "Maszyny",
                "Wybierz maszynę do zmiany statusu.",
                parent=root,
            )
            return
        _open_status_change_dialog(machine)

    def _on_edit() -> None:
        nonlocal rows_cache
        mid = _selected_id()
        if not mid:
            return
        current = _find_machine(mid)
        if not current:
            return
        if not _is_machine_edit_mode():
            _open_machine_usage_window(current)
            return
        if not _require_machine_edit_mode():
            return

        def commit(upd: Dict) -> None:
            nonlocal rows_cache
            upd.setdefault("nr_hali", "1")
            if "zadania" not in upd and isinstance(current.get("zadania"), list):
                upd["zadania"] = current["zadania"]
            new_rows = upsert_machine(rows_cache, upd)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()

        MachineEditDialog(container, row=current, on_ok=commit)

    def _on_del() -> None:
        nonlocal rows_cache
        if not _require_machine_edit_mode():
            return
        mid = _selected_id()
        if not mid:
            return
        if messagebox.askyesno("Usuń", f"Czy usunąć maszynę: {mid}?"):
            new_rows = delete_machine(rows_cache, mid)
            persisted = _save_rows(new_rows)
            rows_cache = list(persisted)
            _on_rows_changed()

    def _on_save() -> None:
        nonlocal rows_cache
        if not _require_machine_edit_mode():
            return
        for row in rows_cache:
            if "nr_hali" not in row or row.get("nr_hali") in (None, ""):
                row["nr_hali"] = "1"
        persisted = _save_rows(rows_cache)
        rows_cache = list(persisted)
        _on_rows_changed()
        info.set(f"Zapisano {len(rows_cache)} maszyn.")

    def _on_tree_select(_event=None) -> None:
        _set_selected_machine(_selected_id())

    btn_import.configure(command=_do_import)
    btn_add.configure(command=_on_add)
    btn_change_status.configure(command=_on_change_status)
    btn_edit.configure(command=_on_edit)
    btn_del.configure(command=_on_del)
    btn_save.configure(command=_on_save)
    btn_mark_done.configure(command=_mark_done)
    btn_restore.configure(command=_restore_plan)
    btn_assign_card.configure(command=_assign_card)
    btn_open_card.configure(command=_open_selected_card)
    btn_clear_search.configure(command=_clear_search)

    tree.bind("<<TreeviewSelect>>", _on_tree_select)
    tree.bind("<Double-1>", lambda _e: _on_edit())
    upcoming_tree.bind("<<TreeviewSelect>>", _on_upcoming_select)
    history_tree.bind("<<TreeviewSelect>>", _on_history_select)
    filter_box.bind("<<ComboboxSelected>>", _apply_filter)
    entry_search.bind("<KeyRelease>", lambda _event: _apply_search())
    entry_search.bind("<Return>", _on_search_enter)
    machine_mode_var.trace_add("write", _refresh_machine_mode_ui)

    _refresh_machine_mode_ui()
    _refresh_schedule_info()
    _recompute_visible_rows()
    _refresh_tree()
    if not initial_machine_id:
        try:
            root.after_idle(_show_review_notice_once)
        except Exception:
            pass
    initial_machine = _find_machine(initial_machine_id)
    if initial_machine is None:
        _populate_details(None)
    else:
        selected_id = str(
            initial_machine.get("id") or initial_machine.get("nr_ewid") or ""
        ).strip()
        _set_selected_machine(selected_id)
        _refresh_tree()
        _open_machine_usage_window(initial_machine)

    logger.info("[Maszyny] Panel otwarty; rekordów: %d", len(rows_cache))
    return tree


def open_machine_usage(
    master: tk.Misc, machine_id: str, *, label: str = ""
) -> tk.Toplevel:
    """Otwórz panel maszyn oraz użytkowanie wskazanej maszyny."""
    win = tk.Toplevel(master)
    win.title(f"Użytkowanie maszyny – {label or machine_id}")
    win.geometry("1200x800")
    win.resizable(True, True)

    container = ttk.Frame(win)
    container.pack(fill="both", expand=True)
    _open_machines_panel(win, container, initial_machine_id=machine_id)
    return win


def panel_maszyny(root, frame, login=None, rola=None):
    try:
        from gui_panel import wm_set_module_source
        from config_manager import ConfigManager, get_machines_path

        cfg = {}
        try:
            cfg = ConfigManager().load()
        except Exception:
            cfg = {}
        wm_set_module_source(root, "Maszyny", get_machines_path(cfg))
    except Exception:
        pass

    for child in frame.winfo_children():
        child.destroy()

    module_frame = ttk.Frame(frame)
    module_frame.pack(fill="both", expand=True)

    _open_maszyny = _open_machines_panel  # alias nazwy
    toolbar = ttk.Frame(module_frame)
    toolbar.pack(fill="x", padx=6, pady=(6, 0))
    target = root
    if hasattr(root, "winfo_toplevel"):
        try:
            target = root.winfo_toplevel()
        except Exception:
            target = root
    ttk.Button(
        toolbar,
        text="Nowa dyspozycja…",
        command=lambda: _maybe_open_dyspo(
            target,
            {
                "typ_dyspozycji": "maszyna",
                "modul_zrodlowy": "maszyny",
            },
        ),
    ).pack(side=tk.RIGHT)

    panel_container = ttk.Frame(module_frame)
    panel_container.pack(fill="both", expand=True)
    _open_maszyny(root, panel_container, Renderer=None)


def init_maszyny_view(
    parent: tk.Misc,
    lista_maszyn: Optional[List[Dict[str, object]]] = None,
    logout_cb: Optional[Callable[[], None]] = None,
    quit_cb: Optional[Callable[[], None]] = None,
    reset_cb: Optional[Callable[[], None]] = None,
) -> MachinesView:
    """Zainicjalizuj widok maszyn w trybie uproszczonym."""

    cfg = Settings(path="config.json", project_root=__file__)
    bg_path = cfg.path_assets("hala.png")
    view = MachinesView(
        parent,
        cfg,
        bg_path=bg_path,
        logout_cb=logout_cb,
        quit_cb=quit_cb,
        reset_cb=reset_cb,
    )
    view.set_records(lista_maszyn or [])
    return view


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Warsztat Menager — Maszyny")
    ensure_theme_applied(root)
    main = tk.Frame(root)
    main.pack(fill="both", expand=True)
    _open_machines_panel(root, main, Renderer=None)
    root.mainloop()
