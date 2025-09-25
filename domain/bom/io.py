"""Obsługa wejścia/wyjścia danych BOM z wykorzystaniem ustawień ścieżek."""

from __future__ import annotations

import csv
import json
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

from config.paths import get_path


def _resolve_path() -> Path:
    """Zwróć ścieżkę do pliku BOM wynikającą z ustawień."""

    raw_path = get_path("bom.file")
    if not raw_path:
        raise FileNotFoundError("Brak skonfigurowanej ścieżki do pliku BOM")
    path = Path(raw_path)
    if not path.suffix:
        # Jeśli w ustawieniach brakuje rozszerzenia, zakładamy JSON jako domyślne.
        path = path.with_suffix(".json")
    return path


def _ensure_parent(path: Path) -> None:
    if path.parent and not path.parent.exists():
        path.parent.mkdir(parents=True, exist_ok=True)


def _load_json(path: Path) -> Any:
    try:
        with path.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    except FileNotFoundError:
        return []


def _load_csv(path: Path) -> list[dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            return [dict(row) for row in reader]
    except FileNotFoundError:
        return []


def _load_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - zależne od środowiska
        raise RuntimeError("Brak obsługi plików XLSX: openpyxl nie jest dostępny") from exc

    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
    except FileNotFoundError:
        return []

    sheet = workbook.active
    rows = list(sheet.rows)
    if not rows:
        return []

    header = [str(cell.value).strip() if cell.value is not None else "" for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        record: dict[str, Any] = {}
        empty = True
        for key, cell in zip(header, row):
            if not key:
                continue
            value = cell.value
            if value is not None and value != "":
                empty = False
            record[key] = value
        if not empty:
            records.append(record)
    return records


def _iter_records(data: Any) -> list[Mapping[str, Any]]:
    if data is None:
        return []
    if isinstance(data, Mapping):
        return [data]
    if isinstance(data, Sequence) and not isinstance(data, (str, bytes, bytearray)):
        out: list[Mapping[str, Any]] = []
        for item in data:
            if isinstance(item, Mapping):
                out.append(item)
            else:
                out.append({"value": item})
        return out
    if isinstance(data, Iterable) and not isinstance(data, (str, bytes, bytearray)):
        out = []
        for item in data:
            if isinstance(item, Mapping):
                out.append(item)
            else:
                out.append({"value": item})
        return out
    return []


def _collect_headers(records: Iterable[Mapping[str, Any]]) -> list[str]:
    headers: list[str] = []
    for rec in records:
        for key in rec.keys():
            if key not in headers:
                headers.append(key)
    return headers


def _save_json(path: Path, data: Any) -> None:
    _ensure_parent(path)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2)


def _save_csv(path: Path, data: Any) -> None:
    records = _iter_records(data)
    headers = _collect_headers(records)
    _ensure_parent(path)
    with path.open("w", encoding="utf-8", newline="") as handle:
        if not headers:
            handle.write("")
            return
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for rec in records:
            writer.writerow({key: rec.get(key, "") for key in headers})


def _save_xlsx(path: Path, data: Any) -> None:
    records = _iter_records(data)
    headers = _collect_headers(records)
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - zależne od środowiska
        raise RuntimeError("Brak obsługi plików XLSX: openpyxl nie jest dostępny") from exc

    _ensure_parent(path)
    workbook = Workbook()
    sheet = workbook.active

    if headers:
        sheet.append(headers)
        for rec in records:
            sheet.append([rec.get(key) for key in headers])
    workbook.save(path)


def bom_load() -> Any:
    """Wczytaj strukturę BOM z pliku określonego w ustawieniach."""

    path = _resolve_path()
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _load_xlsx(path)
    raise ValueError(f"Nieobsługiwane rozszerzenie pliku BOM: {path.suffix}")


def bom_save(data: Any) -> None:
    """Zapisz dane BOM do pliku wynikającego z ustawień."""

    path = _resolve_path()
    suffix = path.suffix.lower()
    if suffix == ".json":
        _save_json(path, data)
        return
    if suffix == ".csv":
        _save_csv(path, data)
        return
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        _save_xlsx(path, data)
        return
    raise ValueError(f"Nieobsługiwane rozszerzenie pliku BOM: {path.suffix}")
