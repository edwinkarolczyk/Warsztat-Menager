"""Independent Excel production plan monitoring engine.

This module deliberately does not import any Warsztat Menager code. It can be
copied together with ``gui.py`` and packaged as a standalone application.
"""

from __future__ import annotations

import csv
import json
import logging
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

APP_DIR = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
CONFIG_PATH = APP_DIR / "config.json"
SNAPSHOT_PATH = APP_DIR / "snapshots" / "current_snapshot.json"
HISTORY_PATH = APP_DIR / "reports" / "history.jsonl"
LOG_PATH = APP_DIR / "logs" / "plan_monitor.log"
SUPPORTED_EXTENSIONS = {".xls", ".xlsx", ".xlsm"}
FIELDS = ("order", "symbol", "quantity", "date", "process")
DEFAULT_CONFIG: dict[str, Any] = {
    "plan_file": "",
    "check_interval_seconds": 60,
    "department_keywords": [],
    "sheet_name": None,
    "header_scan_rows": 15,
    "data_start_row": None,
    "data_end_row": None,
    "column_mapping": {field: "" for field in FIELDS},
}
COLUMN_ALIASES = {
    "order": ("nr zlec", "nr zlec.", "nr zlecenia", "numer zlecenia",
              "zlecenie", "order", "nr"),
    "symbol": ("symbol", "opis", "nazwa", "wyrób", "wyrob", "detal",
               "pozycja", "asortyment", "produkt"),
    "quantity": ("ilość", "ilosc", "ilość szt", "szt", "szt.",
                 "quantity"),
    "date": ("data wysyłki", "data wysylki", "data", "termin", "wysyłka",
             "wysylka", "deadline"),
    "process": ("proces", "status", "operacja", "process"),
}
DEFAULT_COLUMN_MAPPING = {
    "order": "A",
    "symbol": "B",
    "quantity": "C",
    "date": "D",
    "process": "E",
}
HEADER_VALUES = {
    alias for aliases in COLUMN_ALIASES.values() for alias in aliases
} | {"tydzień", "tydzien", "pon", "wt", "śr", "sr", "czw", "pt", "sob",
     "nd"}
CHANGE_LABELS = {
    "new": "NOWE",
    "removed": "USUNIĘTE",
    "quantity_changed": "ILOŚĆ",
    "date_changed": "TERMIN",
    "process_changed": "PROCES",
}


def ensure_directories(base_dir: Path = APP_DIR) -> None:
    """Create all runtime directories required by Plan Monitor."""
    for name in ("snapshots", "reports", "logs", "data"):
        (base_dir / name).mkdir(parents=True, exist_ok=True)


def setup_logging(log_path: Path = LOG_PATH) -> None:
    """Configure simple UTF-8 file logging."""
    ensure_directories(log_path.parent.parent)
    if not any(
        isinstance(handler, logging.FileHandler)
        and Path(handler.baseFilename) == log_path
        for handler in logging.getLogger().handlers
    ):
        handler = logging.FileHandler(log_path, encoding="utf-8")
        handler.setFormatter(
            logging.Formatter("%(asctime)s %(levelname)s %(message)s")
        )
        logging.getLogger().addHandler(handler)
        logging.getLogger().setLevel(logging.INFO)


def load_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    """Load configuration, filling newly introduced options with defaults."""
    config = json.loads(json.dumps(DEFAULT_CONFIG))
    if not path.exists():
        return config
    with path.open("r", encoding="utf-8") as file:
        loaded = json.load(file)
    config.update(loaded)
    mapping = dict(DEFAULT_CONFIG["column_mapping"])
    mapping.update(loaded.get("column_mapping", {}))
    if "deadline" in mapping and not mapping.get("date"):
        mapping["date"] = mapping["deadline"]
    mapping.pop("deadline", None)
    config["column_mapping"] = mapping
    return config


def save_config(config: dict[str, Any], path: Path = CONFIG_PATH) -> None:
    """Persist configuration as human-readable UTF-8 JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(config, file, ensure_ascii=False, indent=2)
        file.write("\n")


def _is_empty(value: Any) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def normalize_text(value: Any) -> str:
    """Return whitespace-normalized text while retaining Polish characters."""
    if _is_empty(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_order(value: Any) -> str:
    """Return a stable order number without a spreadsheet ``.0`` suffix."""
    return normalize_text(value)


def normalize_quantity(value: Any) -> int | float | None:
    """Parse a quantity such as ``450 szt``; return ``None`` if unavailable."""
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"[-+]?\d[\d\s]*(?:[.,]\d+)?", text)
    if not match:
        return None
    number = float(match.group(0).replace(" ", "").replace(",", "."))
    return int(number) if number.is_integer() else number


def safe_date(value: Any) -> str:
    """Format a spreadsheet date safely, including empty pandas values."""
    if _is_empty(value):
        return ""
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.strftime("%Y-%m-%d")
    return normalize_text(value)


def normalize_date(value: Any) -> str:
    """Normalize real Excel dates and preserve textual dates such as ``10 cze``."""
    return safe_date(value)


# Backward-compatible name used by older callers.
normalize_deadline = normalize_date


def normalize_process(value: Any) -> str:
    """Return a whitespace-normalized process description."""
    return normalize_text(value)


def normalize_header(value: Any) -> str:
    """Normalize a possible spreadsheet header for alias matching."""
    return normalize_text(value).lower()


def excel_column_index(reference: Any) -> int | None:
    """Convert an Excel column reference such as ``B`` or ``AA`` to an index."""
    normalized = normalize_text(reference).upper()
    if not re.fullmatch(r"[A-Z]+", normalized):
        return None
    index = 0
    for character in normalized:
        index = index * 26 + ord(character) - ord("A") + 1
    return index - 1


def _column_letter(reference: Any) -> str | None:
    index = excel_column_index(reference)
    return get_column_letter(index + 1) if index is not None else None


def resolve_mapping(
    columns: Iterable[Any], configured: dict[str, str] | None = None
) -> dict[str, str]:
    """Resolve a one-row header iterable for backward-compatible callers."""
    columns = list(columns)
    configured = configured or {}
    rows = [columns]
    return detect_column_mapping(rows, configured)


def detect_column_mapping(
    header_rows: Iterable[Iterable[Any]],
    configured: dict[str, str] | None = None,
) -> dict[str, str]:
    """Detect column letters across multiple header rows, honoring overrides."""
    configured = configured or {}
    detected: dict[str, str] = {}
    for field_name in FIELDS:
        manual = configured.get(field_name, "")
        if field_name == "date":
            manual = manual or configured.get("deadline", "")
        letter = _column_letter(manual)
        if letter:
            detected[field_name] = letter

    for values in header_rows:
        for column_index, value in enumerate(values, start=1):
            normalized = normalize_header(value)
            if not normalized:
                continue
            for field_name, aliases in COLUMN_ALIASES.items():
                if field_name not in detected and normalized in aliases:
                    detected[field_name] = get_column_letter(column_index)

    for field_name, letter in DEFAULT_COLUMN_MAPPING.items():
        detected.setdefault(field_name, letter)
    return detected


def _is_header_symbol(symbol: str) -> bool:
    return normalize_header(symbol).rstrip(":") in HEADER_VALUES


def _records_from_mapping(records: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    return list(records.values())


@dataclass
class ParserDiagnostics:
    """Facts reported after parsing a worksheet."""

    sheet: str = ""
    rows_scanned: int = 0
    records_count: int = 0
    column_mapping: dict[str, str] = field(default_factory=dict)
    skipped_empty: int = 0
    skipped_header: int = 0
    skipped_without_symbol: int = 0
    skipped_non_production: int = 0
    no_order: int = 0
    sample: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class ParseResult:
    """Normalized records together with parser diagnostics."""

    rows: list[dict[str, Any]]
    diagnostics: ParserDiagnostics


def _log_diagnostics(diagnostics: ParserDiagnostics) -> None:
    mapping = " ".join(
        f"{field}={diagnostics.column_mapping.get(field, '-')}"
        for field in FIELDS
    )
    logging.info("[PLAN][PARSE] sheet=%s rows_scanned=%s", diagnostics.sheet,
                 diagnostics.rows_scanned)
    logging.info("[PLAN][PARSE] mapping %s", mapping)
    logging.info(
        "[PLAN][PARSE] records=%s skipped_empty=%s skipped_header=%s "
        "skipped_without_symbol=%s skipped_non_production=%s no_order=%s",
        diagnostics.records_count, diagnostics.skipped_empty,
        diagnostics.skipped_header, diagnostics.skipped_without_symbol,
        diagnostics.skipped_non_production, diagnostics.no_order,
    )
    for row in diagnostics.sample:
        logging.info("[PLAN][PARSE] sample: %s | %s | %s | %s", row["order"],
                     row["symbol"], row["quantity"], row["date"])


def _openpyxl_plan(path: Path, config: dict[str, Any]) -> ParseResult:
    workbook = load_workbook(path, read_only=True, data_only=True)
    requested_sheet = normalize_text(config.get("sheet_name"))
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise ValueError(f"Nie znaleziono arkusza: {requested_sheet}")
        worksheet = workbook[requested_sheet]
    else:
        worksheet = workbook.active or workbook.worksheets[0]

    max_row = worksheet.max_row
    scan_rows = max(1, int(config.get("header_scan_rows") or 15))
    header_values = list(
        worksheet.iter_rows(min_row=1, max_row=min(scan_rows, max_row),
                            values_only=True)
    )
    mapping = detect_column_mapping(header_values, config.get("column_mapping"))
    indexes = {field: excel_column_index(letter) for field, letter in mapping.items()}
    start_row = max(1, int(config.get("data_start_row") or 1))
    end_row = min(max_row, int(config.get("data_end_row") or max_row))
    diagnostics = ParserDiagnostics(sheet=worksheet.title, column_mapping=mapping)
    rows: list[dict[str, Any]] = []
    last_order = ""
    last_date = ""

    for values in worksheet.iter_rows(min_row=start_row, max_row=end_row,
                                      values_only=True):
        diagnostics.rows_scanned += 1
        values = tuple(values)
        if not any(normalize_text(value) for value in values):
            diagnostics.skipped_empty += 1
            continue

        def value_for(field: str) -> Any:
            index = indexes.get(field)
            return values[index] if index is not None and index < len(values) else None

        explicit_order = normalize_order(value_for("order"))
        symbol = normalize_text(value_for("symbol"))
        quantity = normalize_quantity(value_for("quantity"))
        explicit_date = normalize_date(value_for("date"))
        process = normalize_process(value_for("process"))

        if not symbol:
            diagnostics.skipped_without_symbol += 1
            continue
        if _is_header_symbol(symbol):
            diagnostics.skipped_header += 1
            continue
        if len(symbol) < 3:
            diagnostics.skipped_non_production += 1
            continue

        order = explicit_order or last_order
        if quantity is None and not order:
            diagnostics.skipped_non_production += 1
            continue
        if explicit_order and explicit_order != last_order:
            last_date = ""
        row_date = explicit_date or last_date
        row = {
            "order": order,
            "symbol": symbol,
            "quantity": quantity,
            "date": row_date,
            "process": process,
        }
        rows.append(row)
        if explicit_order:
            last_order = explicit_order
        if explicit_date:
            last_date = explicit_date
        if not order:
            diagnostics.no_order += 1

    workbook.close()
    diagnostics.records_count = len(rows)
    diagnostics.sample = rows[:5]
    _log_diagnostics(diagnostics)
    return ParseResult(rows, diagnostics)


def _pandas_xls_plan(path: Path, config: dict[str, Any]) -> ParseResult:
    """Fallback parser for legacy XLS files where openpyxl cannot be used."""
    sheet_name = config.get("sheet_name") or 0
    frame = pd.read_excel(path, sheet_name=sheet_name, header=None)
    values = frame.where(pd.notna(frame), None).values.tolist()
    scan_rows = max(1, int(config.get("header_scan_rows") or 15))
    mapping = detect_column_mapping(values[:scan_rows], config.get("column_mapping"))
    indexes = {field: excel_column_index(letter) for field, letter in mapping.items()}
    start_row = max(1, int(config.get("data_start_row") or 1))
    end_row = min(len(values), int(config.get("data_end_row") or len(values)))
    diagnostics = ParserDiagnostics(sheet=str(sheet_name), column_mapping=mapping)
    rows: list[dict[str, Any]] = []
    last_order = ""
    last_date = ""
    for values_row in values[start_row - 1:end_row]:
        diagnostics.rows_scanned += 1
        if not any(normalize_text(value) for value in values_row):
            diagnostics.skipped_empty += 1
            continue
        def get(field_name: str) -> Any:
            index = indexes[field_name]
            return values_row[index] if index < len(values_row) else None
        explicit_order = normalize_order(get("order"))
        symbol = normalize_text(get("symbol"))
        quantity = normalize_quantity(get("quantity"))
        explicit_date = normalize_date(get("date"))
        process = normalize_process(get("process"))
        if not symbol:
            diagnostics.skipped_without_symbol += 1
            continue
        if _is_header_symbol(symbol):
            diagnostics.skipped_header += 1
            continue
        if len(symbol) < 3:
            diagnostics.skipped_non_production += 1
            continue
        order = explicit_order or last_order
        if quantity is None and not order:
            diagnostics.skipped_non_production += 1
            continue
        if explicit_order and explicit_order != last_order:
            last_date = ""
        row = {"order": order, "symbol": symbol, "quantity": quantity,
               "date": explicit_date or last_date, "process": process}
        rows.append(row)
        last_order = explicit_order or last_order
        last_date = explicit_date or last_date
        if not order:
            diagnostics.no_order += 1
    diagnostics.records_count = len(rows)
    diagnostics.sample = rows[:5]
    _log_diagnostics(diagnostics)
    return ParseResult(rows, diagnostics)


def parse_plan(path: str | Path, config: dict[str, Any] | None = None) -> ParseResult:
    """Parse the complete selected worksheet and return diagnostics."""
    plan_path = Path(path)
    if plan_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError("Obsługiwane formaty planu: xls, xlsx, xlsm.")
    options = json.loads(json.dumps(DEFAULT_CONFIG))
    if config:
        options.update(config)
        options["column_mapping"] = config.get("column_mapping", config)
    if plan_path.suffix.lower() == ".xls":
        return _pandas_xls_plan(plan_path, options)
    return _openpyxl_plan(plan_path, options)


def read_plan(path: str | Path, mapping: dict[str, str] | None = None,
              **options: Any) -> list[dict[str, Any]]:
    """Backward-compatible convenience wrapper returning normalized rows only."""
    config = dict(options)
    config["column_mapping"] = mapping or {}
    return parse_plan(path, config).rows


def position_key(row: dict[str, Any]) -> str:
    """Build a stable order+symbol key, falling back to symbol+date."""
    order = normalize_order(row.get("order"))
    symbol = normalize_text(row.get("symbol"))
    row_date = normalize_date(row.get("date", row.get("deadline")))
    if order:
        return f"{order}|{symbol}"
    return f"NO_ORDER|{symbol}|{row_date}"


def concerns_department(row: dict[str, Any], keywords: Iterable[str]) -> bool:
    """Check symbol/description against configured department keywords."""
    haystack = f'{row.get("symbol", "")} {row.get("order", "")}'.upper()
    return any(keyword.strip().upper() in haystack for keyword in keywords
               if keyword.strip())


def _change(change_type: str, row: dict[str, Any], old: Any, new: Any,
            keywords: Iterable[str], timestamp: str) -> dict[str, Any]:
    return {"timestamp": timestamp, "type": change_type,
            "order": row.get("order", ""), "symbol": row.get("symbol", ""),
            "old": old, "new": new,
            "department_related": concerns_department(row, keywords)}


def _canonical_row(row: dict[str, Any]) -> dict[str, Any]:
    canonical = dict(row)
    canonical["date"] = normalize_date(row.get("date", row.get("deadline")))
    canonical.pop("deadline", None)
    canonical.setdefault("process", "")
    return canonical


def compare_plans(previous: list[dict[str, Any]], current: list[dict[str, Any]],
                  keywords: Iterable[str] = (), timestamp: str | None = None
                  ) -> list[dict[str, Any]]:
    """Return new, removed and field-level changes between complete snapshots."""
    timestamp = timestamp or datetime.now().isoformat(timespec="seconds")
    old_by_key = {position_key(row): _canonical_row(row) for row in previous}
    new_by_key = {position_key(row): _canonical_row(row) for row in current}
    changes: list[dict[str, Any]] = []
    for key in sorted(old_by_key.keys() & new_by_key.keys()):
        _append_field_changes(changes, old_by_key[key], new_by_key[key],
                              keywords, timestamp)
    for key in sorted(new_by_key.keys() - old_by_key.keys()):
        row = new_by_key[key]
        changes.append(_change("new", row, "", row, keywords, timestamp))
    for key in sorted(old_by_key.keys() - new_by_key.keys()):
        row = old_by_key[key]
        changes.append(_change("removed", row, row, "", keywords, timestamp))
    return changes


def _append_field_changes(changes: list[dict[str, Any]],
                          old_row: dict[str, Any], new_row: dict[str, Any],
                          keywords: Iterable[str], timestamp: str) -> None:
    for field_name, change_type in (("quantity", "quantity_changed"),
                                    ("date", "date_changed"),
                                    ("process", "process_changed")):
        if old_row.get(field_name) != new_row.get(field_name):
            changes.append(_change(change_type, new_row, old_row.get(field_name),
                                   new_row.get(field_name), keywords, timestamp))


def load_snapshot(path: Path = SNAPSHOT_PATH) -> dict[str, Any] | None:
    """Read the previous snapshot, if one exists."""
    if not path.exists():
        return None
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def snapshot_rows(snapshot: dict[str, Any] | None) -> list[dict[str, Any]]:
    """Load records from current snapshots and the legacy rows-list format."""
    if not snapshot:
        return []
    if isinstance(snapshot.get("records"), dict):
        return _records_from_mapping(snapshot["records"])
    return snapshot.get("rows", [])


def save_snapshot(rows: list[dict[str, Any]], metadata: dict[str, Any],
                  path: Path = SNAPSHOT_PATH) -> None:
    """Persist all normalized records and parser metadata as UTF-8 JSON."""
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "created_at": metadata.get("read_at", datetime.now().isoformat(timespec="seconds")),
        "source_file": metadata.get("plan_file", ""),
        "sheet": metadata.get("sheet", ""),
        "parser": metadata.get("parser", {}),
        "records": {position_key(row): row for row in rows},
    }
    with path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2)
        file.write("\n")


def append_history(changes: Iterable[dict[str, Any]], path: Path = HISTORY_PATH) -> None:
    """Append one JSON object per detected change."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as file:
        for change in changes:
            file.write(json.dumps(change, ensure_ascii=False) + "\n")


def load_history(path: Path = HISTORY_PATH) -> list[dict[str, Any]]:
    """Load the JSONL report history."""
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8") as file:
        return [json.loads(line) for line in file if line.strip()]


def format_parser_summary(parser: dict[str, Any] | ParserDiagnostics | None) -> list[str]:
    """Return report lines describing the latest worksheet parse."""
    if isinstance(parser, ParserDiagnostics):
        parser = asdict(parser)
    parser = parser or {}
    mapping = parser.get("column_mapping", {})
    columns = "/".join(mapping.get(field, "-") for field in FIELDS)
    return [f'Przeskanowano wierszy: {parser.get("rows_scanned", 0)}',
            f'Znaleziono pozycji: {parser.get("records_count", 0)}',
            f"Użyte kolumny: {columns}"]


def export_changes(changes: list[dict[str, Any]], path: str | Path,
                   parser: dict[str, Any] | ParserDiagnostics | None = None) -> None:
    """Export parser summary and visible changes to TXT or CSV."""
    export_path = Path(path)
    rows = [display_row(change) for change in changes]
    headers = ["Data", "Typ zmiany", "Nr zlecenia", "Symbol",
               "Stara wartość", "Nowa wartość", "Dotyczy działu"]
    summary = format_parser_summary(parser)
    if export_path.suffix.lower() == ".csv":
        with export_path.open("w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file, delimiter=";")
            for line in summary:
                writer.writerow([line])
            writer.writerow([])
            writer.writerow(headers)
            writer.writerows(rows)
    elif export_path.suffix.lower() == ".txt":
        with export_path.open("w", encoding="utf-8") as file:
            file.write("\n".join(summary) + "\n\n")
            file.write(" | ".join(headers) + "\n")
            file.write("-" * 120 + "\n")
            for row in rows:
                file.write(" | ".join(map(str, row)) + "\n")
    else:
        raise ValueError("Raport można eksportować wyłącznie do TXT lub CSV.")


def display_row(change: dict[str, Any]) -> tuple[Any, ...]:
    """Convert a history entry into GUI/export table columns."""
    old, new = change.get("old", ""), change.get("new", "")
    if isinstance(old, dict):
        old = old.get("quantity", "")
    if isinstance(new, dict):
        new = new.get("quantity", "")
    return (change.get("timestamp", ""),
            CHANGE_LABELS.get(change.get("type", ""), change.get("type", "")),
            change.get("order", ""), change.get("symbol", ""), old, new,
            "DOTYCZY DZIAŁU" if change.get("department_related") else "POZA DZIAŁEM")


@dataclass
class CheckResult:
    """Result passed safely from the worker thread to the GUI thread."""

    status: str
    checked_at: str
    file_modified_at: str = ""
    changes: list[dict[str, Any]] | None = None
    message: str = ""
    parser: dict[str, Any] = field(default_factory=dict)
    rows: list[dict[str, Any]] = field(default_factory=list)
    error: str = ""


class PlanMonitor:
    """Stateful, GUI-independent production plan checker."""

    def __init__(self, config_path: Path = CONFIG_PATH,
                 snapshot_path: Path = SNAPSHOT_PATH,
                 history_path: Path = HISTORY_PATH,
                 dispositions_path: Path | None = None,
                 reader: Callable[..., Any] | None = None) -> None:
        self.config_path = config_path
        self.snapshot_path = snapshot_path
        self.history_path = history_path
        self.reader = reader
        self.config = load_config(config_path)
        self.last_signature: tuple[float, int] | None = None
        self.last_parser: dict[str, Any] = {}
        self.last_rows: list[dict[str, Any]] = []

    def reload_config(self) -> None:
        """Reload settings edited in the GUI."""
        self.config = load_config(self.config_path)

    def _read(self, plan_file: str) -> ParseResult:
        if self.reader is None:
            return parse_plan(plan_file, self.config)
        result = self.reader(plan_file, self.config.get("column_mapping", {}))
        if isinstance(result, ParseResult):
            return result
        rows = list(result)
        diagnostics = ParserDiagnostics(records_count=len(rows), sample=rows[:5])
        return ParseResult(rows, diagnostics)

    def check(self, force: bool = False) -> CheckResult:
        """Check metadata and process a changed source file without raising."""
        checked_at = datetime.now().strftime("%H:%M:%S")
        plan_file = self.config.get("plan_file", "")
        if not plan_file:
            return CheckResult("error", checked_at, message="Nie wybrano pliku planu.",
                               error="Nie wybrano pliku planu.")
        try:
            stat = os.stat(plan_file)
            signature = (stat.st_mtime, stat.st_size)
            modified_at = datetime.fromtimestamp(stat.st_mtime).strftime("%H:%M:%S")
            snapshot = load_snapshot(self.snapshot_path)
            if not force and snapshot and signature == self.last_signature:
                return CheckResult("unchanged", checked_at, modified_at, [],
                                   "Plik nie zmienił się od ostatniego odczytu.",
                                   self.last_parser, self.last_rows)
            parsed = self._read(plan_file)
            rows = parsed.rows
            parser = asdict(parsed.diagnostics)
            previous = snapshot_rows(snapshot)
            changes = compare_plans(previous, rows,
                                    self.config.get("department_keywords", []))
            metadata = {"plan_file": plan_file,
                        "read_at": datetime.now().isoformat(timespec="seconds"),
                        "sheet": parsed.diagnostics.sheet, "parser": parser}
            save_snapshot(rows, metadata, self.snapshot_path)
            if changes:
                append_history(changes, self.history_path)
            self.last_signature = signature
            self.last_parser = parser
            self.last_rows = rows
            logging.info("Odczytano plan: %s; wykryto zmian: %s", plan_file,
                         len(changes))
            return CheckResult("changed" if changes else "unchanged", checked_at,
                               modified_at, changes, notification_message(changes),
                               parser, rows)
        except Exception as error:  # Keep network failures away from the GUI.
            logging.exception("Nie udało się odczytać planu: %s", plan_file)
            message = f"Błąd odczytu pliku: {error}"
            return CheckResult("error", checked_at, message=message, error=message)


def notification_message(changes: list[dict[str, Any]]) -> str:
    """Build the requested summary notification."""
    if not changes:
        return "Brak zmian w planie"
    related = sum(bool(change.get("department_related")) for change in changes)
    return f"Wykryto {len(changes)} zmiany. {related} dotyczą Twojego działu."


def main() -> None:
    """Start the Tk GUI lazily, keeping backend imports headless-friendly."""
    setup_logging()
    try:
        from .gui import run
    except ImportError:
        from gui import run
    run()


if __name__ == "__main__":
    main()
